import streamlit as st
import pandas as pd
import os
import plotly.express as px
import json
import datetime
from zoneinfo import ZoneInfo
import base64
import requests
import io

# Importar lógica modularizada
import core_logic

# ---------------------------------------------------
# ✅ CONSTANTES GLOBALES
# ---------------------------------------------------
TZ_CO = ZoneInfo("America/Bogota")
TZ_UTC = ZoneInfo("UTC")
GRUPOS_OPERATIVOS = ["INSP-CALDAS", "INSP-RIS"]
CODIGOS_ADICIONALES = ["12163", "12164", "10793", "12170", "10842", "10772", "10445"]

# Mapeo maestro de inspectores a supervisores
SUPERVISORES_DICT = {k.upper(): v for k, v in {
    "ARIZA MARIN SERGIO": "ANDRES ARROYAVE", "ANDRES ARROYAVE": "ANDRES ARROYAVE",
    "BEDOYA DIEGO ALEJANDRO": "DANNY DE LA CRUZ", "DANNY DE LA CRUZ": "DANNY DE LA CRUZ",
    "CARVAJAL RESTREPO JUAN DAVID": "JANIER MARIN", "JANIER MARIN": "JANIER MARIN",
     "ECHEVERRY CARDONA JHON STIVEN": "JANIER MARIN", "GALLEGO CADAVID NORBEY": "DANNY DE LA CRUZ",
    "GIRALDO GARCIA SIGIFREDO": "ANDRES ARROYAVE", "LOPEZ PINEDA CESAR AUGUSTO": "JANIER MARIN",
    "NOREÑA GIRALDO GEOVANNY": "ANDRES ARROYAVE", "OSPINA CASTELLANOS ANDERSON": "CRISTIAN CHICA",
    "OSPINA RODRIGUEZ DANIEL ALBERTO": "ANDRES ARROYAVE", "RUIZ DILON MARLON ANDREY": "ANDRES ARROYAVE",
    "LARGO OSORIO JOSE OMAR": "ANDRES ARROYAVE", "PULGARIN QUINTERO JULIAN ANDRES": "DANNY DE LA CRUZ",
    "TAYACK TRUJILLO DEIVER EVELIO": "ANDRES ARROYAVE", "PATIÑO CIFUENTES RICARDO": "JANIER MARIN", "VARGAS FRANCO JHON EDISON": "CRISTIAN CHICA",
    "CARDONA CANO NELSON": "CRISTIAN CHICA", "CARDONA OROZCO JULIAN ANDRES": "ANDRES ARROYAVE",
    "GRISALES CUERVO JUAN DAVID": "JANIER MARIN", "LEON MARIN LEONARDO FABIO": "JANIER MARIN",
    "CARDONA CASTANO DIDIER ORLANDO": "CRISTIAN CHICA",
    "TORRES HERNANDEZ JOHN JAMES": "ANDRES ARROYAVE", "COBO HOYOS JUAN MANUEL": "CRISTIAN CHICA",
    "OSPINA NARANJO BERNARDO": "CRISTIAN CHICA", "COGOLLO FIGUEROA RANDY": "DANNY DE LA CRUZ",
    "ARIAS TORO YEISON": "DANNY DE LA CRUZ", "MIRANDA FRANCO EFRAIN": "DANNY DE LA CRUZ",
    "ARDILA MORA GUSTAVO ADOLFO": "DANNY DE LA CRUZ", "LOPEZ VELEZ ESTEBAN": "JANIER MARIN",
    "GALEANO GRISALEZ RICARDO": "DANNY DE LA CRUZ", "CAICEDO ESCOBAR JUNIOR SANTIAGO": "JANIER MARIN",
    "BUITRAGO RAMIREZ LEONARD": "CRISTIAN CHICA",
    "BORJAS WILLY ALEXANDER": "ANDRES ARROYAVE", "MARIN LEON JAISSON JOAQUIN": "CRISTIAN CHICA",
    "AMAYA HINCAPIE JUAN CARLOS": "CRISTIAN CHICA", "BEDOYA SANCHEZ CRISTIAN DAVID": "ANDRES ARROYAVE",
    "RAMIREZ WILSON ENRIQUE": "CRISTIAN CHICA", "CANO MORALES JIMY ALFREDO": "ANDRES ARROYAVE",
    "CASTRO CASTAÑO JUAN DAVID": "CRISTIAN CHICA",   "VILLA LOAIZA JHEISON ESTIBEN": "CRISTIAN CHICA", "CÁRDENAS GALIANO HAROLD MAURICIO": "JANIER MARIN",
    "VARGAS CORREA VICTOR ALFONSO": "DANNY DE LA CRUZ", "VILLA MERA CHRISTIAN DAVID": "JANIER MARIN",
    "AVENDAÑO GARCIA JUAN NEPOMUCENO": "ANDRES ARROYAVE", "PELAEZ TATIS GABRIEL ESTEBAN": "CRISTIAN CHICA",
    "CHICA RAMIREZ CRISTIAN ALBERTO": "CRISTIAN CHICA"
}.items()}

# ---------------------------------------------------
# ✅ CONFIGURACIÓN GENERAL DEL DASHBOARD
# ---------------------------------------------------
st.set_page_config(
    page_title="DASHBOARD INSPECTORES e&c",
    layout="wide"
)

# ✅ ESTILOS GLOBALES
st.markdown("""
    <style>
    /* Fondo con degradado sutil para un aspecto moderno */
    .stApp { 
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    /* Contenedor principal del dashboard en blanco puro al entrar */
    .stMainBlockContainer {
        background-color: #ffffff;
    }
    /* Estilo de Tarjetas Profesionales para KPIs */
    .metric-card {
        background-color: #ffffff;
        padding: 0.8rem;
        border-radius: 12px;
        box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.1), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
        border-left: 5px solid #2F9331;
        margin-bottom: 0.5rem;
    }
    .metric-label {
        color: #64748b;
        font-size: 0.75rem;
        font-weight: 700;
        text-transform: uppercase;
    }
    .metric-value {
        color: #2F9331;
        font-size: 1.35rem;
        font-weight: 800;
        margin-top: 5px;
    }
    /* Forzar color azul en Pills y controles segmentados seleccionados */
    div[data-testid="stPills"] button[aria-checked="true"],
    div[data-testid="stSegmentedControl"] button[aria-checked="true"],
    div[data-baseweb="tag"] {
        background-color: #39A935 !important;
        border-color: #39A935 !important;
        color: white !important;
    }
    /* Texto blanco en elementos seleccionados */
    div[data-testid="stPills"] button[aria-checked="true"] p,
    div[data-testid="stSegmentedControl"] button[aria-checked="true"] p {
        color: white !important;
    }
    /* Estilos específicos para la tabla de SST */
    .sst-container {
        background-color: #f8fafc;
        padding: 1.5rem;
        border-radius: 15px;
        border: 1px solid #e2e8f0;
        margin-top: 1rem;
        box-shadow: inset 0 2px 4px 0 rgba(0, 0, 0, 0.05);
    }
    </style>
""", unsafe_allow_html=True)

# -------------------------------------------------
# ✅ FUNCIONES DE CACHÉ. (MEJORA DE RENDIMIENTO)
# -------------------------------------------------

@st.cache_data(ttl=600)  # Cache por 10 minutos para datos de GitHub

@st.cache_data(ttl=300)
def get_processed_agendas_data(repo, token):
    df_raw, _ = fetch_github_excel(repo, "BITACORA.xlsx", token)
    if df_raw.empty:
        return pd.DataFrame()
    
    df = df_raw.copy()
    df.columns = [str(c).strip().lower() for c in df.columns]
    
    columnas_req = [
        "grupo", "prioridad", "estado",
        "fecha de visita", "fecha de ejecucion",
        "inspector", "contrato", "direccion",
        "localidad", "detalle de tarea"
    ]
    
    # Verificar columnas
    if not all(c in df.columns for c in columnas_req):
        return pd.DataFrame() # O manejar error
        
    df["grupo"] = df["grupo"].astype(str).str.upper().str.strip()
    df = df[df["grupo"].isin(["INSP-CALDAS", "INSP-RIS"])].copy()
    if "prioridad" in df.columns:
        df = df[df["prioridad"].astype(str).str.upper().str.strip() == "ALTA"].copy()
    
    df["fecha de visita"] = pd.to_datetime(df["fecha de visita"], errors="coerce")
    df["fecha de ejecucion"] = pd.to_datetime(df["fecha de ejecucion"], errors="coerce")
    
    ahora_colombia = datetime.datetime.now(ZoneInfo("America/Bogota")).replace(tzinfo=None)
    
    # Cruce con Programación
    df_prog_aux, _ = fetch_github_excel(repo, "PROGRAMACION.xlsx", token)
    if not df_prog_aux.empty:
        df_prog_aux.columns = [str(c).strip().lower() for c in df_prog_aux.columns]
        if "contrato" in df_prog_aux.columns and "hora agenda" in df_prog_aux.columns:
            df["contrato"] = df["contrato"].astype(str).str.strip()
            df_prog_aux["contrato"] = df_prog_aux["contrato"].astype(str).str.strip()
            df_prog_aux = df_prog_aux.drop_duplicates(subset=["contrato"])
            df = df.merge(df_prog_aux[["contrato", "hora agenda"]], on="contrato", how="left")
            
    def calc_alerta(row):
        visita = row["fecha de visita"]
        if pd.isna(visita) or visita > ahora_colombia: return "OK"
        ha = str(row.get("hora agenda", "")).upper()
        if any(txt in ha for txt in ["TRANSCURSO DE LA MAÑANA", "TRANSCURSO DE LA TARDE", "TRANSCURSO DEL DIA", "TRANSCURSO DEL DÍA", "JORNADA MAÑANA", "JORNADA TARDE", "JORNADA DE MAÑANA"]):
            return "OK"
        return "ALERTA"
        
    df["estado_alerta"] = df.apply(calc_alerta, axis=1)
    return df

def fetch_github_excel(repo, path, token, branch="main"):
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        data = r.json()
        sha = data.get("sha")
        try:
            # Usar download_url es más confiable para archivos binarios y evita el límite de 1MB de la API
            download_url = data.get("download_url")
            if download_url:
                resp = requests.get(download_url, headers=headers)
                content = resp.content
            elif "content" in data:
                content = base64.b64decode(data["content"])
            else:
                return pd.DataFrame(), sha
            
            # Dejamos que pandas detecte el motor (engine) automáticamente para soportar .xls y .xlsx
            return pd.read_excel(io.BytesIO(content)), sha
        except Exception as e:
            st.error(f"❌ Error al procesar el Excel '{path}' desde GitHub: {e}")
            return pd.DataFrame(), sha
    return pd.DataFrame(), None

@st.cache_data(ttl=300)
def fetch_github_json(repo, path, token):
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    r = requests.get(url, headers=headers)
    if r.status_code == 200:
        try:
            data = r.json()
            content = base64.b64decode(data["content"]).decode("utf-8")
            return json.loads(content), data.get("sha")
        except Exception:
            return {}, None
    return {}, None

def save_github_json(repo, path, token, data, message, branch="main"):
    """Guarda un diccionario o lista como JSON en GitHub."""
    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{repo}/contents/{path}"
    
    # Obtener el SHA actual para permitir el reemplazo
    r = requests.get(url, headers=headers)
    sha = r.json().get("sha") if r.status_code == 200 else None
    
    content_b64 = base64.b64encode(json.dumps(data, indent=2, ensure_ascii=False).encode("utf-8")).decode("utf-8")
    payload = {"message": message, "content": content_b64, "branch": branch}
    
    if sha:
        payload["sha"] = sha
        
    return requests.put(url, headers=headers, json=payload)

@st.cache_data(ttl=600)
def load_local_bitacora(path):
    if os.path.exists(path):
        try:
            df = pd.read_excel(path)
        except Exception:
            return None
            
        df.columns = [str(c).strip().lower() for c in df.columns]
        
        # Pre-procesamiento de nombres e inspectores
        if "inspector" in df.columns:
            df["inspector"] = df["inspector"].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)
            
        # Mapeo de supervisores usando la constante global
        df["supervisor"] = df["inspector"].map(SUPERVISORES_DICT).fillna("SIN SUPERVISOR")
        
        # Conversión de Fechas y Horas una sola vez
        if "fecha de ejecucion" in df.columns:
            df["fecha"] = pd.to_datetime(df["fecha de ejecucion"], errors="coerce").dt.date
            
        # Parseo de horas (simplificado)
        for col in ["hora inicio", "hora inicio de recorrido", "hora final"]:
            if col in df.columns:
                df[col + "_parsed"] = pd.to_datetime(df[col].astype(str), errors='coerce').dt.time

        if "tiempo de tarea" in df.columns:
            df["tiempo_tarea_td"] = pd.to_timedelta(df["tiempo de tarea"].astype(str), errors="coerce")

        # --- PRE-CÁLCULOS GLOBALES PARA RENDIMIENTO ---
        import datetime
        
        # 1. Efectiva
        valores_efectivos = [
            "INSPECCIONADA",
            "INSPECCIONADA CON DEFECTO NO CRITICO",
            "INSPECCIONADA CON DEFECTO CRITICO",
            "CERTIFICADA",
            "CERTIFICADA CON NOVEDAD"
        ]
        if "cierre" in df.columns:
            df["efectiva"] = df["cierre"].isin(valores_efectivos)
        else:
            df["efectiva"] = False
            
        # 2. Tiempos (solo si existen las columnas parseadas)
        if "hora inicio_parsed" in df.columns and "hora inicio de recorrido_parsed" in df.columns:
            def calc_recorrido(row):
                hi = row.get("hora inicio_parsed")
                hr = row.get("hora inicio de recorrido_parsed")
                if not isinstance(hi, datetime.time) or not isinstance(hr, datetime.time):
                    return pd.NaT
                dt_hi = datetime.datetime.combine(datetime.date.today(), hi)
                dt_hr = datetime.datetime.combine(datetime.date.today(), hr)
                return dt_hi - dt_hr if dt_hi >= dt_hr else pd.NaT
            df["tiempo_recorrido_td"] = df.apply(calc_recorrido, axis=1)
        else:
            df["tiempo_recorrido_td"] = pd.NaT
            
        # 3. Decimales (útil para promedios rápidos)
        def to_dec(h):
            if not isinstance(h, datetime.time): return None
            return h.hour + h.minute / 60.0 + h.second / 3600.0
            
        if "hora inicio_parsed" in df.columns:
            df["ini_dec_tmp"] = df["hora inicio_parsed"].apply(to_dec)
        if "hora final_parsed" in df.columns:
            df["fin_dec_tmp"] = df["hora final_parsed"].apply(to_dec)
            
        # 4. Puntualidad (minutos tarde)
        hora_oficial = datetime.time(7, 30)
        def calc_tarde(h):
            if not isinstance(h, datetime.time): return None
            h1 = datetime.datetime.combine(datetime.date.today(), h)
            h2 = datetime.datetime.combine(datetime.date.today(), hora_oficial)
            return int((h1 - h2).total_seconds() / 60)
            
        if "hora inicio_parsed" in df.columns:
            df["minutos_tarde"] = df["hora inicio_parsed"].apply(calc_tarde)
            def calc_estado(m):
                if m is None or pd.isna(m): return "SIN INICIO"
                if m <= 0: return "Puntual"
                if m <= 15: return "Tarde"
                return "Muy tarde"
            df["estado_puntualidad"] = df["minutos_tarde"].apply(calc_estado)

        return df
    return None


@st.cache_data(ttl=300)
def process_uploaded_mensual_file(file_content):
    import io
    # file_content is bytes to avoid hashing issues with UploadedFile
    df_m = pd.read_excel(io.BytesIO(file_content))
    df_m.columns = [str(c).strip().lower() for c in df_m.columns]
    
    if "inspector" in df_m.columns:
        df_m["inspector"] = df_m["inspector"].astype(str).str.upper().str.strip().str.replace(r"\s+", " ", regex=True)
    df_m["supervisor"] = df_m["inspector"].map(SUPERVISORES_DICT).fillna("SIN SUPERVISOR")
    
    if "fecha de ejecucion" in df_m.columns:
        df_m["fecha"] = pd.to_datetime(df_m["fecha de ejecucion"], errors="coerce").dt.date
        
    for col in ["hora inicio", "hora inicio de recorrido", "hora final"]:
        if col in df_m.columns:
            df_m[col + "_parsed"] = pd.to_datetime(df_m[col].astype(str), errors='coerce').dt.time

    if "tiempo de tarea" in df_m.columns:
        df_m["tiempo_tarea_td"] = pd.to_timedelta(df_m["tiempo de tarea"].astype(str), errors="coerce")
        
    return df_m


@st.cache_data(ttl=600)
def process_sst_data(df_sst):
    if df_sst.empty:
        return pd.DataFrame()
        
    inspectores_sst = df_sst["inspector"].unique()
    df_result = pd.DataFrame({"inspector": inspectores_sst})
    
    # 1. Preoperacionales
    df_preoperacional = df_sst[
        df_sst["tipo de trabajo"].str.upper().str.contains("INSPECCION DE VEHICULO", na=False)
    ].copy()
    
    # 2. Ausentismo
    df_ausentismo = df_sst[
        df_sst["tipo de trabajo"].str.upper().str.contains(
            "|".join(["AUSENTISMO", "VACACIONES", "INCAPACIDAD", "PERMISO REMUNERADO", "PERMISO NO REMUNERADO", "SUSPENSION", "LICENCIA LUTO", "LICENCIA DE PATERNIDAD", "LICENCIA NO REMUNERADA", "PERMISO SINDICAL", "SANCIÓN"]),
            na=False
        )
    ].copy()
    
    # 3. Operacional Final
    ausentismos_tipos = [
        "AUSENTISMO", "VACACIONES", "INCAPACIDAD", "PERMISO REMUNERADO",
        "PERMISO NO REMUNERADO", "SUSPENSION", "LICENCIA LUTO",
        "LICENCIA DE PATERNIDAD", "LICENCIA NO REMUNERADA",
        "PERMISO SINDICAL", "SANCIÓN"
    ]
    df_operacional_final = df_sst[
        (~df_sst["tipo de trabajo"].str.upper().str.contains("INSPECCION DE VEHICULO", na=False)) &
        (~df_sst["tipo de trabajo"].str.upper().str.contains("|".join(ausentismos_tipos), na=False)) &
        (df_sst["tiempo de tarea"].notna())
    ].copy()
    
    df_preoperacional_agg = df_preoperacional.groupby("inspector")["tiempo_tarea_td"].sum().reset_index(name="HORA_PREOPERACIONAL")
    df_operacional_final_agg = df_operacional_final.groupby("inspector")["hora_final_parsed"].max().reset_index(name="HORA_OPERACIONAL_FINAL")
    df_ausentismo_agg = df_ausentismo.groupby("inspector")["tiempo_tarea_td"].sum().reset_index(name="TIEMPO_AUSENTISMO")
    
    df_result = df_result.merge(df_preoperacional_agg, on="inspector", how="left")
    df_result = df_result.merge(df_operacional_final_agg, on="inspector", how="left")
    df_result = df_result.merge(df_ausentismo_agg, on="inspector", how="left")
    
    def format_timedelta(td):
        if pd.isna(td): return "NO TIENE"
        total_seconds = int(td.total_seconds())
        hours, remainder = divmod(total_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{hours:02}:{minutes:02}:{seconds:02}"
        
    def format_time(t):
        if pd.isna(t) or t is None: return "NO TIENE"
        return t.strftime("%H:%M:%S")

    df_result["HORA PREOPERACIONAL"] = df_result["HORA_PREOPERACIONAL"].apply(format_timedelta)
    df_result["HORA OPERACIONAL FINAL"] = df_result["HORA_OPERACIONAL_FINAL"].apply(format_time)
    df_result["AUSENTISMO"] = df_result["TIEMPO_AUSENTISMO"].apply(format_timedelta)
    
    df_result.rename(columns={"inspector": "INSPECTOR"}, inplace=True)
    return df_result[["INSPECTOR", "HORA PREOPERACIONAL", "HORA OPERACIONAL FINAL", "AUSENTISMO"]]

@st.cache_data(ttl=600)
def process_adicionales_data(df):
    """Procesa los datos de programación de forma cacheada para evitar lentitud en filtros."""
    if df.empty: return df
    df.columns = [str(c).strip().lower() for c in df.columns]
    
    # Inicializamos la columna para evitar KeyError si no se encuentra una fecha válida
    df["dias de asignacion"] = 0

    # Filtro de códigos
    if "codigo_tipo_trabajo" in df.columns:
        df = df[df["codigo_tipo_trabajo"].astype(str).isin(CODIGOS_ADICIONALES)]
        
    # Cálculo de fechas (heurística).
    posibles_fechas = [
        "fecha de asignacion", "fecha asignacion", "asignacion", "fecha", 
        "fecha cargue", "fecha programacion", "f_asignacion", "fecha_asignacion"
    ]
    col_fecha = next((c for c in df.columns if c in posibles_fechas), None)
    
    if col_fecha:
        df[col_fecha] = pd.to_datetime(df[col_fecha], errors="coerce")
        hoy = datetime.datetime.now(TZ_CO).date()
        df["dias de asignacion"] = df[col_fecha].apply(lambda x: (hoy - x.date()).days if pd.notna(x) else 0)
    
    return df

@st.cache_data(ttl=600)
@st.cache_data(ttl=600)
def extract_excel_links(path):
    from openpyxl import load_workbook
    if not os.path.exists(path): return pd.DataFrame()
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    try:
        c_ins, c_fac, c_vp = headers.index("inspector")+1, headers.index("foto de fachada")+1, headers.index("foto de vp")+1
    except ValueError: return pd.DataFrame()
    
    links = []
    for row in ws.iter_rows(min_row=2):
        links.append({
            "inspector": row[c_ins-1].value,
            "link_fachada": row[c_fac-1].hyperlink.target if row[c_fac-1].hyperlink else None,
            "link_vp": row[c_vp-1].hyperlink.target if row[c_vp-1].hyperlink else None
        })
    return pd.DataFrame(links)


def parse_tiempo_tarea(valor):
    try:
        return pd.to_timedelta(str(valor))
    except Exception:
        return pd.NaT

def hora_to_decimal(h):
    if h is None or h == "SIN HORA" or pd.isna(h):
        return None
    if isinstance(h, str): return None
    return h.hour + h.minute / 60 + h.second / 3600

def decimal_to_hora(d):
    if d is None or pd.isna(d):
        return None
    h = int(d)
    m = int((d - h) * 60)
    return datetime.time(h, m)

def hora_to_string(h):
    return h.strftime("%I:%M %p") if h and not isinstance(h, str) else "—"

def td_to_str(td):
    if pd.isna(td):
        return "—"
    s = int(td.total_seconds())
    h = s // 3600
    m = (s % 3600) // 60
    s2 = s % 60
    return f"{h}h {m}m {s2}s" if h > 0 else f"{m}m {s2}s"

def render_kpi(label, value, icon=""):
    st.markdown(f"""
        <div class="metric-card">
            <div class="metric-label">{icon} {label}</div>
            <div class="metric-value">{value}</div>
        </div>
    """, unsafe_allow_html=True)

@st.dialog("Detalle de la Tarea")
def mostrar_detalle_tarea(contrato, detalle, direccion="—", fecha="—", localidad="—", inspector="—"):
    st.markdown(f"### 📋 Contrato: {contrato}")
    st.markdown("---")
    c1, c2 = st.columns(2)
    with c1:
        st.markdown(f"**🏠 Dirección:** {direccion}")
        st.markdown(f"**📍 Localidad:** {localidad}")
    with c2:
        st.markdown(f"**📅 Fecha Visita:** {fecha}")
        st.markdown(f"**👤 Inspector:** {inspector}")
    
    st.markdown("**📝 Detalle de la tarea:**")
    st.info(detalle if detalle else "Sin detalle adicional.")

def color_estado(val):
    if val == "Puntual": return 'background-color: #d4edda; color: #155724;'
    if "tarde" in str(val).lower(): return 'background-color: #fff3cd; color: #856404;'
    return ''

# -------------------------------------------------
# ZONA HORARIA COLOMBIA
# -------------------------------------------------



if "usuario" not in st.session_state:
    st.session_state.usuario = None
    st.session_state.rol = None

# --- LÓGICA DE CIERRE DE SESIÓN POR INACTIVIDAD (5 MINUTOS) ---
if "last_activity" not in st.session_state:
    st.session_state.last_activity = datetime.datetime.now()

if st.session_state.usuario is not None:
    ahora = datetime.datetime.now()
    segundos_inactivo = (ahora - st.session_state.last_activity).total_seconds()
    
    if segundos_inactivo > 1800:  # 1800 segundos = 30 minutos
        st.session_state.usuario = None
        st.session_state.rol = None
        st.warning("⚠️ Sesión cerrada por inactividad (30 minutos).")
        st.rerun()
    st.session_state.last_activity = ahora

def cargar_usuarios():
    if os.path.exists("USUARIOS.json"):
        with open("USUARIOS.json", "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

# ===================================================
# ✅ INTERFAZ DE INICIO DE SESIÓN (WIDE CARD CORPORATIVO)
# ===================================================
if st.session_state.usuario is None:
    import base64
    import os
    
    logo_data = ""
    if os.path.exists("logo.png"):
        with open("logo.png", "rb") as f:
            logo_data = base64.b64encode(f.read()).decode("utf-8")
            
    # Intentar cargar el imagotipo si el usuario lo subió
    imagotipo_data = ""
    if os.path.exists("E&C IMAGOTIPO VER VERDE.png"):
        with open("E&C IMAGOTIPO VER VERDE.png", "rb") as f:
            imagotipo_data = base64.b64encode(f.read()).decode("utf-8")
    else:
        imagotipo_data = logo_data  # Fallback si no está

    st.markdown(f"""
        <style>
        /* Ocultar el top-bar de Streamlit */
        [data-testid="stHeader"] {{ display: none !important; }}
        
        /* Fondo premium de pantalla completa */
        .stApp {{
            background: linear-gradient(135deg, #e2e8f0 0%, #cbd5e1 100%) !important;
        }}
        
        /* Convertir el contenedor principal en flex para centrar la tarjeta */
        [data-testid="stMainBlockContainer"] {{
            display: flex !important;
            justify-content: center !important;
            align-items: center !important;
            min-height: 100vh !important;
            padding: 2rem !important;
        }}
        
        /* Nuestra tarjeta principal basada en el stHorizontalBlock */
        div[data-testid="stHorizontalBlock"] {{
            background-color: transparent !important;
            border-radius: 20px !important;
            box-shadow: 0 25px 50px -12px rgba(0, 0, 0, 0.25) !important;
            overflow: hidden !important;
            width: 95% !important;
            max-width: 1600px !important; /* Tarjeta gigantesca (ocupa casi toda la pantalla) */
            margin: auto !important;
            display: flex !important;
            align-items: stretch !important;
            position: relative;
        }}

        /* Mitad Izquierda (Formulario) -> CORRECTO SELECTOR DE STREAMLIT */
        div[data-testid="stColumn"]:nth-of-type(1) {{
            background-color: white !important;
            padding: 5rem 6rem 5rem 4rem !important; /* Más padding a la derecha para que el logo no choque con el formulario */
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
            margin: 0 !important;
        }}

        /* Mitad Derecha (Texto verde) -> CORRECTO SELECTOR DE STREAMLIT */
        div[data-testid="stColumn"]:nth-of-type(2) {{
            background: linear-gradient(135deg, #1A3820 0%, #39A935 100%) !important;
            padding: 5rem 3rem 5rem 8rem !important; /* El 8rem de la izquierda actúa como un escudo protector para el logo */
            display: flex !important;
            flex-direction: column !important;
            justify-content: center !important;
            align-items: center !important; /* Centrar todo el contenido para evitar choque con el logo */
            color: white !important;
            margin: 0 !important;
        }}

        /* Textos Derecha */
        .right-title {{
            font-size: 3.5rem;
            font-weight: 800;
            line-height: 1.1;
            margin-bottom: 1.5rem;
            color: white;
            font-family: sans-serif;
            position: relative;
            z-index: 5;
            text-align: center;
        }}
        .right-text {{
            font-size: 1.1rem;
            line-height: 1.6;
            color: rgba(255,255,255,0.95);
            text-align: center;
            font-family: sans-serif;
            margin-bottom: 2rem;
            position: relative;
            z-index: 5;
        }}
        .chat-bubble {{
            background: white;
            color: #1A3820;
            padding: 1rem 1.5rem;
            border-radius: 20px 20px 20px 0px;
            display: inline-block;
            font-weight: bold;
            box-shadow: 0 10px 20px rgba(0,0,0,0.15);
            font-size: 1.05rem;
            position: relative;
            z-index: 5;
            width: fit-content;
        }}

        /* Logo Completo (Formulario Izquierdo) */
        .login-logo {{
            text-align: center;
            margin-bottom: 1.5rem;
        }}
        .login-logo img {{
            width: 180px;
        }}
        
        /* Círculo central superpuesto con el Imagotipo */
        .circle-logo {{
            position: absolute;
            left: -100px; /* Más desplazado a la izquierda según solicitud */
            top: 50%;
            transform: translateY(-50%);
            width: 160px; /* Logo más grande */
            height: 160px;
            background: white;
            border-radius: 50%;
            box-shadow: 0 10px 30px rgba(0,0,0,0.15);
            display: flex;
            align-items: center;
            justify-content: center;
            z-index: 20;
            padding: 20px;
        }}
        .circle-logo img {{
            width: 100%;
            height: 100%;
            object-fit: contain;
        }}
        </style>
    """, unsafe_allow_html=True)

    col_form, col_info = st.columns([1, 1], gap="small")

    with col_form:
        # Logo completo encima del formulario (como lo pidió el usuario)
        img_src = f"data:image/png;base64,{logo_data}" if logo_data else ""
        if img_src:
            st.markdown(f"""
                <div class="login-logo">
                    <img src="{img_src}" alt="e&c Logo">
                </div>
            """, unsafe_allow_html=True)
            
        st.markdown("<h3 style='color: #2F9331; text-align: center; font-weight: 800; margin-bottom: 1.5rem; letter-spacing: 1px;'>DASHBOARD EJE CAFETERO</h3>", unsafe_allow_html=True)
        
        usuarios = cargar_usuarios()
        usuario_input = st.text_input("Usuario", placeholder="Ej: Juan Perez")
        pin_input = st.text_input("PIN de seguridad", type="password", max_chars=4, placeholder="****")

        st.write("")
        if st.button("🚀 INGRESAR AL SISTEMA", use_container_width=True, type="primary"):
            if usuario_input in usuarios and pin_input == usuarios[usuario_input]["pin"]:
                st.session_state.usuario = usuario_input
                st.session_state.rol = usuarios[usuario_input]["rol"]
                st.empty()
                st.rerun()
            else:
                st.error("❌ Usuario o PIN incorrectos. Intenta de nuevo.")

    with col_info:
        # Imagotipo en el círculo central superpuesto (como en GoDoWorks)
        imago_src = f"data:image/png;base64,{imagotipo_data}" if imagotipo_data else ""
        
        # Fecha y hora dinámica en español
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        ahora_co = datetime.datetime.now(TZ_CO)
        fecha_str = f"{ahora_co.day} de {meses[ahora_co.month - 1]} del {ahora_co.year}"
        hora_str = ahora_co.strftime("%I:%M %p")
        
        st.markdown(f"""
            <div class="circle-logo">
                <img src="{imago_src}" alt="e&c Imagotipo">
            </div>
            <div class="right-title" style="font-size: 2.8rem; line-height: 1.3;">SEGUIMIENTO OPERATIVO<br>EJE CAFETERO</div>
            <div class="right-text" style="margin-top: 1rem; margin-bottom: 2.5rem;">
                <div style="font-size: 4.5rem; font-weight: 900; letter-spacing: 2px; color: white;">{hora_str}</div>
                <div style="font-size: 1.6rem; font-weight: 400; opacity: 0.9; margin-top: 0.5rem; text-transform: uppercase;">{fecha_str}</div>
            </div>
            <div style="display: flex; justify-content: center; width: 100%;">
                <div class="chat-bubble">
                    <span style='color:#39A935;'>¡Hola!</span> ¿Necesita Ayuda?
                </div>
            </div>
        """, unsafe_allow_html=True)

    st.stop()

# -------------------------------------------------
# BOTÓN CERRAR SESIÓN
# -------------------------------------------------
col_vacio, col_logout = st.columns([8, 1])

with col_logout:
    if st.button("🚪 Cerrar sesión"):
        st.session_state.usuario = None
        st.session_state.rol = None
        st.rerun()

# ===================================================
# ✅ MOSTRAR METADATA DE ACTUALIZACIÓN (GLOBAL)
# ===================================================
def obtener_texto_meta(info_dict):
    if not info_dict or "ultima_actualizacion" not in info_dict:
        return "—", "—"
    try:
        fecha_utc = datetime.datetime.strptime(
            info_dict.get("ultima_actualizacion"), "%Y-%m-%d %H:%M:%S"
        ).replace(tzinfo=TZ_UTC)
        fecha_col = fecha_utc.astimezone(TZ_CO)
        return fecha_col.strftime("%Y-%m-%d %H:%M:%S"), info_dict.get("usuario_actualizo", "—")
    except:
        return "—", "—"

token_meta = st.secrets["github"]["token"]
repo_meta = st.secrets["github"]["repo"]

# Leer info de ambos archivos desde GitHub para consistencia global
info_bitacora_meta, _ = fetch_github_json(repo_meta, "BITACORA_INFO.json", token_meta)
info_programacion_meta, _ = fetch_github_json(repo_meta, "PROGRAMACION_INFO.json", token_meta)

f_bit, u_bit = obtener_texto_meta(info_bitacora_meta)
f_prog, u_prog = obtener_texto_meta(info_programacion_meta)

st.markdown(
    f"""
    <div style='display: flex; justify-content: space-between; padding: 0px 15px; margin-bottom: -15px;'>
        <div style='color: #64748b; font-size: 0.8rem; font-family: sans-serif;'>
            🕓 <b>Bitácora:</b> {f_bit} | 👤 {u_bit}
        </div>
        <div style='color: #64748b; font-size: 0.8rem; font-family: sans-serif; text-align: right;'>
            📅 <b>Programación:</b> {f_prog} | 👤 {u_prog}
        </div>
    </div>
    <hr style='margin: 10px 0px; border-top: 1px solid #e2e8f0;'>
    """, 
    unsafe_allow_html=True
)

# ---------------------------------------------------

# ---------------------------------------------------
# ✅ LISTA MAESTRA DE INSPECTORES
# ---------------------------------------------------
inspectores_lista = sorted([
  
    "ARIZA MARIN SERGIO",
    "ANDRES ARROYAVE",
    "BEDOYA DIEGO ALEJANDRO",
    "DANNY DE LA CRUZ",
    "CARVAJAL RESTREPO JUAN DAVID",
    "JANIER MARIN",
    "CHAVARRIAGA JUAN MANUEL",
    "CRISTIAN CHICA",
    "ECHEVERRY CARDONA JHON STIVEN",
    "GALLEGO CADAVID NORBEY",
    "GIRALDO GARCIA SIGIFREDO",
    "LOPEZ PINEDA CESAR AUGUSTO",
    "NOREÑA GIRALDO GEOVANNY",
    "OSPINA CASTELLANOS ANDERSON",
    "OSPINA RODRIGUEZ DANIEL ALBERTO",
    "RUIZ DILON MARLON ANDREY",
    "LARGO OSORIO JOSE OMAR",
    "PULGARIN QUINTERO JULIAN ANDRES",
    "TAYACK TRUJILLO DEIVER EVELIO",
    "RUIZ ARENAS JUAN CAMILO",
    "PATIÑO CIFUENTES RICARDO",
    "VARGAS FRANCO JHON EDISON",
    "CARDONA CANO NELSON",
    "CARDONA OROZCO JULIAN ANDRES",
    "GRISALES CUERVO JUAN DAVID",
    "LEON MARIN LEONARDO FABIO",
    "VELASQUEZ TAPASCO JHON DIEGO",
    "CARDONA CASTANO DIDIER ORLANDO",
    "TORRES HERNANDEZ JOHN JAMES",
    "COBO HOYOS JUAN MANUEL",
    "OSPINA NARANJO BERNARDO",
    "COGOLLO FIGUEROA RANDY",
    "ARIAS TORO YEISON",
    "MIRANDA FRANCO EFRAIN",
    "ARDILA MORA GUSTAVO ADOLFO",
    "LOPEZ VELEZ ESTEBAN",
    "GALEANO GRISALEZ RICARDO",
    "CAICEDO ESCOBAR JUNIOR SANTIAGO",
    "OTERO CAICEDO ANYEMBER",
    "BUITRAGO RAMIREZ LEONARD",
    "BORJAS WILLY ALEXANDER",
    "MARIN LEON JAISSON JOAQUIN",
    "AMAYA HINCAPIE JUAN CARLOS",
    "BEDOYA SANCHEZ CRISTIAN DAVID",
    "RAMIREZ WILSON ENRIQUE",
    "CANO MORALES JIMY ALFREDO",
    "CASTRO CASTAÑO JUAN DAVID",
    "LOAIZA GAMBA JHON ALEXANDER",
    "VILLA LOAIZA JHEISON ESTIBEN",
    "CÁRDENAS GALIANO HAROLD MAURICIO",
    "VARGAS CORREA VICTOR ALFONSO",
    "VILLA MERA CHRISTIAN DAVID",
    "AVENDAÑO GARCIA JUAN NEPOMUCENO",
    "PELAEZ TATIS GABRIEL ESTEBAN",
    "CHICA RAMIREZ CRISTIAN ALBERTO",
])
# ---------------------------------------------------
# HEADER CON LOGO A LA DERECHA (CORRECTO)
# -------------------------------------------------
col_titulo, col_logo = st.columns([8, 2])

with col_titulo:
    st.markdown(
        "<h1 style='margin-bottom: 0;'>📊 DASHBOARD INSPECTORES E&C</h1>",
        unsafe_allow_html=True
    )

with col_logo:
    st.image(
        "logo.png",
        use_container_width=True,
        caption=""
    )
# ---------------------------------------------------

# ===================================================
# CARGA ÚNICA DE BITÁCORA (BASE GLOBAL)
# ===================================================
with st.spinner("🔄 Sincronizando datos con el servidor... Un momento por favor"):
    archivo_bitacora = "BITACORA.xlsx"
    
    # Carga base de datos
    df_bitacora_base = load_local_bitacora(archivo_bitacora)
    
    if df_bitacora_base is None:
        st.error("❌ No se encontró el archivo BITACORA.xlsx.")
        st.stop()
    
    # Extraer links una sola vez aquí para evitar lentitud en el Tab 2
    df_links_global = extract_excel_links(archivo_bitacora)

# ✅ CREAR PESTAÑAS
# ---------------------------------------------------
tab_operacion, tab_inv_v2, tab_sst, tab_subir = st.tabs([
    "🚀 OPERACIÓN", 
    "📦 INVENTARIO", 
    "🛡️ SST", 
    "📤 CARGAR DATOS"
])

with tab_operacion:
    tab_diario, tab_mensual, tab_agendas, tab_adicionales, tab_asignadas = st.tabs([
            "🕒 Seguimiento Diario", 
            "📅 Seguimiento Mensual", 
            "📅 Seguimiento agendas", 
            "🏭 SEGUIMIENTO ADICIONALES", 
            "📌 Órdenes Asignadas"
        ])

# ===================================================
    with tab_diario:
        st.subheader("🕒 Control Operativo e&c")
        st.subheader("Eje Cafetero")
    
        # Usar la base ya cargada y procesada
        df_bitacora = df_bitacora_base.copy()
    
        if df_bitacora is None:
            st.error(
                "❌ Error al procesar la bitácora."
            )
            st.stop()
    
        # -------------------------------------------------
        # ✅ EXCLUIR GRUPOS NO OPERATIVOS
        # -------------------------------------------------
        if "grupo" in df_bitacora.columns:
            df_bitacora["grupo"] = (
                df_bitacora["grupo"]
                .astype(str)
                .str.upper()
                .str.strip()
            )
    
            # ✅ Restaurada exclusión de grupos no operativos y administrativos
            grupos_no_operativos = ["SST-NAL", "SUPERVISIONES", "SUSP-ANT", "ADMINISTRATIVO", "SUSPENSIONES", "ADMIN"]
    
            # Filtrar por grupo y también excluir cierres administrativos de los KPIs
            if "cierre" in df_bitacora.columns:
                df_bitacora = df_bitacora[~df_bitacora["cierre"].astype(str).str.upper().str.contains("ADMINISTRATIVO", na=False)]
    
    
            df_bitacora = df_bitacora[
                ~df_bitacora["grupo"].isin(grupos_no_operativos)
            ]
    
        if df_bitacora.empty:
            st.warning(
                "⚠️ No hay datos disponibles después del filtro por GRUPO.\n"
                "Esto indica que el archivo solo contiene grupos no operativos."
            )
            st.stop()
    
        # Renombrar columnas parseadas para lógica existente
        df_bitacora["hora_inicio"] = df_bitacora["hora inicio_parsed"].fillna("SIN HORA")
        df_bitacora["hora_inicio_recorrido"] = df_bitacora["hora inicio de recorrido_parsed"]
        df_bitacora["hora_final"] = df_bitacora["hora final_parsed"]
    
        # -------------------------------------------
        # -------------------------------------------
        # FILTRO DE FECHA
        # -------------------------------------------
        # -------------------------------------------
        # ✅ PANEL DE FILTROS COMPACTO (TIPO BOX)
        # -------------------------------------------
        with st.container(border=True):
            col_f1, col_f2, col_f3 = st.columns([1, 1.2, 1.2])
            
            with col_f1:
                fechas_validas = sorted(df_bitacora["fecha"].dropna().unique())
                fecha_sel = st.selectbox("📅 Fecha de consulta:", fechas_validas)
                # Datos base para la fecha seleccionada
                df_base_fecha = df_bitacora[df_bitacora["fecha"] == fecha_sel].copy()
    
            # Opciones estables para la fecha elegida
            opc_sups = sorted(df_base_fecha["supervisor"].unique())
            opc_insps = sorted(df_base_fecha["inspector"].unique())
    
            with col_f2:
                supervisores_sel = st.pills("👥 Supervisores:", opc_sups, selection_mode="multi", default=opc_sups, key=f"pills_sup_{fecha_sel}")
    
            with col_f3:
                with st.popover("🔍 Seleccionar Inspectores", use_container_width=True):
                    inspectores_sel = st.multiselect("Filtro de inspectores:", opc_insps, default=opc_insps, key=f"ms_insp_{fecha_sel}")
    
            # Aplicación del filtro final
            if not supervisores_sel or not inspectores_sel:
                st.warning("⚠️ Selecciona al menos un supervisor e inspector para ver los datos.")
                st.stop()
                
            df2 = df_base_fecha[(df_base_fecha["supervisor"].isin(supervisores_sel)) & (df_base_fecha["inspector"].isin(inspectores_sel))].copy()
    
        # -------------------------------------------
        # ⏱️ TIEMPO DE RECORRIDO (Calculado sobre datos filtrados)
        # -------------------------------------------
        pass # Calculado en load_local_bitacora
    
    # ===================================================
      # ===================================================
        # ✅ TAB 2 — PARTE 4 / 5
        # Agrupación diaria, puntualidad y estado
        # ===================================================
    
        # ---------------------------------------------------
        # AGRUPACIÓN DIARIA POR INSPECTOR (ESTABLE)
        # ---------------------------------------------------
        primeras = (
            df2.sort_values("hora_inicio")
            .groupby("inspector", as_index=False)
            .first()[["inspector", "hora_inicio", "localidad", "supervisor", "minutos_tarde", "estado_puntualidad"]]
        )
        primeras.rename(columns={"estado_puntualidad": "estado"}, inplace=True)
    
        ultimas = (
            df2.sort_values("hora_final")
            .groupby("inspector", as_index=False)
            .last()[["inspector", "hora_final"]]
        )
    
        df_agrupado = primeras.merge(
            ultimas,
            on="inspector",
            how="left"
        )
    
        # ---------------------------------------------------
        # PUNTUALIDAD (usa SOLO la primera hora del día)
        # ---------------------------------------------------
        pass # Calculado en load_local_bitacora e integrado en df_agrupado
    
        # ---------------------------------------------------
        # PRODUCCIÓN (MARCAR ÓRDENES EFECTIVAS)
        # ---------------------------------------------------
        pass # Calculado en load_local_bitacora
    
        total_ordenes = df2.shape[0]
        total_efectivas = df2["efectiva"].sum()
    
        porcentaje = (
            round((total_efectivas / total_ordenes) * 100, 1)
            if total_ordenes > 0 else 0
        )
    
        # ---------------------------------------------------
        # ÓRDENES EFECTIVAS CON TIEMPO VÁLIDO
        # ---------------------------------------------------
        df_eff = df2[
            (df2["efectiva"] == True) &
            (df2["tiempo_tarea_td"].notna())
        ]
    
        # ---------------------------------------------------
        # ✅ KPI: PROMEDIO HORA DE INICIO
        # (PRIMERA TAREA DEL DÍA POR INSPECTOR)
        # ---------------------------------------------------
        # Calculamos decimal antes de agrupar para evitar errores con datetime.time en el agg
        pass # Ya existe df2['ini_dec_tmp']
        df_inicio_jornada = df2[df2["ini_dec_tmp"].notna()].groupby("inspector", as_index=False).agg(ini_dec=("ini_dec_tmp", "min"))
    
        prom_ini = df_inicio_jornada["ini_dec"].mean()
    
        hora_prom_ini = (
            hora_to_string(decimal_to_hora(prom_ini))
            if pd.notna(prom_ini) else "—"
        )
    
        # ---------------------------------------------------
        # ✅ KPI: PROMEDIO HORA DE FIN
        # (ÚLTIMA TAREA DEL DÍA POR INSPECTOR)
        # ---------------------------------------------------
        # Calculamos decimal antes de agrupar para evitar errores con datetime.time en el agg
        pass # Ya existe df2['fin_dec_tmp']
        df_fin_jornada = df2[df2["fin_dec_tmp"].notna()].groupby("inspector", as_index=False).agg(fin_dec=("fin_dec_tmp", "max"))
    
        prom_fin = df_fin_jornada["fin_dec"].mean()
    
        hora_prom_fin = (
            hora_to_string(decimal_to_hora(prom_fin))
            if pd.notna(prom_fin) else "—"
        )
    
        # ---------------------------------------------------
        # ✅ KPI: PROMEDIO TIEMPO POR TAREA (SOLO EFECTIVAS)
        # ---------------------------------------------------
        tiempo_prom_str = (
            td_to_str(df_eff["tiempo_tarea_td"].mean())
            if not df_eff.empty else "—"
        )
    
        # ---------------------------------------------------
        # RESUMEN POR INSPECTOR (SOLO PARA CÁLCULO)
        # ---------------------------------------------------
        # ---------------------------------------------------
        # RESUMEN POR INSPECTOR
        # ---------------------------------------------------
        resumen = (
            df2.groupby("inspector")
            .apply(lambda x: pd.Series({
                "total_ordenes": x.shape[0],
                "ordenes_efectivas": x["efectiva"].sum(),
                "porcentaje_efectividad":
                    round((x["efectiva"].sum() / x.shape[0]) * 100, 1)
                    if x.shape[0] > 0 else 0,
                "promedio_tiempo_tarea":
                    td_to_str(
                        x.loc[x["efectiva"], "tiempo_tarea_td"].mean()
                    ),
                "ordenes_sin_recorrido": x["tiempo_recorrido_td"].isna().sum(),
                "promedio_tiempo_recorrido": td_to_str(x["tiempo_recorrido_td"].mean())
            }))
            .reset_index()
        )
    
        df_tabla = df_agrupado.merge(resumen, on="inspector", how="left")
        df_tabla = df_tabla.fillna({
            "hora_inicio": "—",
            "hora_final": "—",
            "localidad": "—",
            "estado": "SIN ACTIVIDAD",
            "total_ordenes": 0,
            "ordenes_efectivas": 0,
            "ordenes_sin_recorrido": 0,
            "porcentaje_efectividad": 0,
            "promedio_tiempo_tarea": "—",
            "promedio_tiempo_recorrido": "—"
        })
    
        # ✅ KPI: PROMEDIO TIEMPO DE RECORRIDO (GENERAL)
        prom_recorrido_global = (
            td_to_str(df2["tiempo_recorrido_td"].mean())
            if not df2["tiempo_recorrido_td"].dropna().empty else "—"
        )
    
    
        # ===================================================
        # ✅ DISEÑO DE PESTAÑA: KPIs IZQUIERDA | TABLA DERECHA
        # ===================================================
        col_kpis, col_main_view = st.columns([1, 4])
    
        with col_kpis:
            st.markdown("### ⭐ KPIs")
            render_kpi("Inicio Prom.", hora_prom_ini, "⏰")
            render_kpi("Fin Prom.", hora_prom_fin, "🕒")
            render_kpi("T. Tarea Prom.", tiempo_prom_str, "🕓")
            render_kpi("Recorrido Prom.", prom_recorrido_global, "🚗")
            render_kpi("Total Tareas", total_ordenes, "📋")
            render_kpi("Efectivas", total_efectivas, "✅")
            render_kpi("% Efectividad", f"{porcentaje}%", "📈")
    
        with col_main_view:
            # ---------------------------------------------------
            # PREPARACIÓN DE LA TABLA ESTILIZADA
            # ---------------------------------------------------
            columnas_tabla = [
                "inspector", "supervisor", "fecha", "hora_inicio", "hora_final",
                "localidad", "estado", "total_ordenes", "ordenes_efectivas",
                "ordenes_sin_recorrido", "porcentaje_efectividad",
                "promedio_tiempo_tarea", "promedio_tiempo_recorrido"
            ]
            columnas_disponibles = [c for c in columnas_tabla if c in df_tabla.columns]
    
            st.markdown("### 📋 Tabla de inspecciones del día")
    
            # Aplicar estilos: Centrar todo menos inspector y aplicar colores a estado
            df_styled = (
                df_tabla[columnas_disponibles]
                .style.set_properties(**{'text-align': 'center'})
                .set_properties(subset=['inspector'], **{'text-align': 'left'})
                .map(color_estado, subset=['estado'])
            )
    
            st.dataframe(
                df_styled,
                use_container_width=True,
                height=645,
                hide_index=True,
                column_config={
                    "porcentaje_efectividad": st.column_config.NumberColumn(
                        "Efectividad %",
                        format="%.1f%%"
                    )
                }
            )
    
        # ---------------------------------------------------
        # 📝 INFORME DE DESEMPEÑO DEL DÍA
        # ---------------------------------------------------
        st.markdown("### 📝 Informe de Desempeño del Día")
        with st.container(border=True):
            h_col1, h_col2 = st.columns(2)
            if not resumen.empty:
                best_eff = resumen.loc[resumen["ordenes_efectivas"].idxmax()]
                worst_eff = resumen.loc[resumen["ordenes_efectivas"].idxmin()]
                most_no_rec = resumen.loc[resumen["ordenes_sin_recorrido"].idxmax()]
                
                with h_col1:
                    st.markdown(f"🏆 **Más órdenes efectivas:** {best_eff['inspector']} ({int(best_eff['ordenes_efectivas'])})")
                    st.markdown(f"📉 **Menos órdenes efectivas:** {worst_eff['inspector']} ({int(worst_eff['ordenes_efectivas'])})")
                    st.markdown(f"🚗 **Más órdenes sin recorrido:** {most_no_rec['inspector']} ({int(most_no_rec['ordenes_sin_recorrido'])})")
    
                df_ini_check = df_agrupado[df_agrupado["hora_inicio"] != "SIN HORA"].copy()
                late_insp, late_val = "—", "—"
                if not df_ini_check.empty:
                    df_ini_check["dec"] = df_ini_check["hora_inicio"].apply(hora_to_decimal)
                    row_late = df_ini_check.loc[df_ini_check["dec"].idxmax()]
                    late_insp, late_val = row_late["inspector"], hora_to_string(row_late["hora_inicio"])
    
                avg_rec_series = df2.groupby("inspector")["tiempo_recorrido_td"].mean()
                max_rec_insp, max_rec_val = "—", "—"
                if not avg_rec_series.dropna().empty:
                    max_rec_insp = avg_rec_series.idxmax()
                    max_rec_val = td_to_str(avg_rec_series.max())
    
                avg_task_series = df2.loc[df2["efectiva"]].groupby("inspector")["tiempo_tarea_td"].mean()
                max_task_insp, max_task_val = "—", "—"
                if not avg_task_series.dropna().empty:
                    max_task_insp = avg_task_series.idxmax()
                    max_task_val = td_to_str(avg_task_series.max())
    
                with h_col2:
                    st.markdown(f"🕒 **Inicio más tarde:** {late_insp} ({late_val})")
                    st.markdown(f"🛣️ **Promedio de recorrido más extenso:** {max_rec_insp} ({max_rec_val})")
                    st.markdown(f"🕓 **Más tiempo promedio por tarea:** {max_task_insp} ({max_task_val})")
    
        # ===================================================
        # 🚨 INSPECTORES SIN ACTIVIDAD EN LA FECHA
        # ===================================================
        st.markdown("### 🚨 Inspectores sin actividad registrada")
    
        inspectores_con_actividad = set(df2["inspector"].str.upper().str.strip().unique())
    
        inspectores_del_filtro = [
            insp for insp in inspectores_lista
            if SUPERVISORES_DICT.get(insp.upper(), "SIN SUPERVISOR") in supervisores_sel
        ]
    
        inspectores_sin_actividad = [
            insp for insp in inspectores_del_filtro
            if insp.upper().strip() not in inspectores_con_actividad
        ]
    
        if inspectores_sin_actividad:
            df_sin_actividad = pd.DataFrame({
                "Inspector": inspectores_sin_actividad
            })
            df_sin_actividad["Supervisor"] = df_sin_actividad["Inspector"].apply(
                lambda x: SUPERVISORES_DICT.get(x.upper(), "SIN SUPERVISOR")
            )
            df_sin_actividad = df_sin_actividad.sort_values("Supervisor")
    
            st.error(f"🚨 {len(inspectores_sin_actividad)} inspector(es) sin actividad registrada para {fecha_sel}")
            st.dataframe(df_sin_actividad, use_container_width=True)
        else:
            st.success("✅ Todos los inspectores tienen actividad registrada para esta fecha.")
    
        # ===================================================
        # 📊 Producción por inspector (órdenes efectivas)
      
        # ===================================================
        st.markdown("## 📊 Producción por inspector (órdenes efectivas)")
    
        df_prod = (
            df2[df2["efectiva"] == True]
            .groupby("inspector")
            .size()
            .reset_index(name="Órdenes efectivas")
            .sort_values("Órdenes efectivas", ascending=True)
        )
    
        if df_prod.empty:
            st.info("⚠️ No hay órdenes efectivas para esta fecha.")
        else:
            # Colores según producción
            def color_por_produccion(valor):
                if valor <= 3:
                    return "#dc3545"      # rojo
                elif valor <= 6:
                    return "#f5b7b1"      # rosado
                elif valor <= 8:
                    return "#f7dc6f"      # amarillo
                else:
                    return "#28a745"      # verde
    
            df_prod["color"] = df_prod["Órdenes efectivas"].apply(color_por_produccion)
    
            fig_prod = px.bar(
                df_prod,
                y="inspector",
                x="Órdenes efectivas",
                orientation="h",
                text="Órdenes efectivas",
                title="Órdenes efectivas por inspector"
            )
    
            # Colores y barras más gruesas
            fig_prod.update_traces(
                marker_color=df_prod["color"],
                textposition="outside",
                textfont_size=28,
                cliponaxis=False
            )
    
            fig_prod.update_layout(
                bargap=0.15,              # barras más gruesas
                xaxis_title="Órdenes efectivas",
                yaxis_title="Inspector",
                font=dict(size=18),
                height=650
            )
    
            st.plotly_chart(fig_prod, use_container_width=True)
     
    # ===================================================
    # ✅ TAB — SEGUIMIENTO MENSUAL
    # ===================================================
    with tab_mensual:
        st.subheader("📅 Consolidado Mensual / Rango de Fechas")
        
        # Opción para subir un archivo independiente para el consolidado mensual
        archivo_mensual = st.file_uploader(
            "📂 Sube una bitácora independiente para el análisis mensual (Rango de varios días)", 
            type=["xlsx", "xls"], 
            key="uploader_mensual"
        )
        
        if archivo_mensual is not None:
            with st.spinner("Procesando archivo mensual..."):
                file_bytes = archivo_mensual.getvalue()
                df_m = process_uploaded_mensual_file(file_bytes)
            st.success("✅ Archivo mensual cargado correctamente.")
        else:
            st.info("👆 Puedes subir un archivo consolidado aquí. Si no, se usará la bitácora diaria actual.")
            df_m = df_bitacora_base.copy()
    
        if "grupo" in df_m.columns:
            df_m["grupo"] = df_m["grupo"].astype(str).str.upper().str.strip()
            grupos_no_operativos = ["SST-NAL", "SUPERVISIONES", "SUSP-ANT", "ADMINISTRATIVO", "SUSPENSIONES", "ADMIN"]
            if "cierre" in df_m.columns:
                df_m = df_m[~df_m["cierre"].astype(str).str.upper().str.contains("ADMINISTRATIVO", na=False)]
            df_m = df_m[~df_m["grupo"].isin(grupos_no_operativos)]
    
        # Renombrar columnas para consistencia
        df_m["hora_inicio"] = df_m["hora inicio_parsed"]
        df_m["hora_final"] = df_m["hora final_parsed"]
        
        # Filtros de Rango
        with st.container(border=True):
            c1, c2, c3 = st.columns([1.5, 1.2, 1.2])
            with c1:
                # Asegurar que hay fechas válidas antes de calcular min/max para evitar errores de tipo
                fechas_validas = df_m["fecha"].dropna()
                if not fechas_validas.empty:
                    min_d, max_d = fechas_validas.min(), fechas_validas.max()
                else:
                    min_d = max_d = datetime.date.today()
                    
                rango_fecha = st.date_input("📅 Seleccionar lapso de días:", value=(min_d, max_d), min_value=min_d, max_value=max_d)
            
            if isinstance(rango_fecha, tuple) and len(rango_fecha) == 2:
                start_date, end_date = rango_fecha
                df_m = df_m[(df_m["fecha"] >= start_date) & (df_m["fecha"] <= end_date)]
            
            opc_sups = sorted(df_m["supervisor"].unique())
            with c2:
                sups_sel = st.pills("👥 Supervisores:", opc_sups, selection_mode="multi", default=opc_sups, key="m_sup_pills")
            with c3:
                opc_insps = sorted(df_m[df_m["supervisor"].isin(sups_sel)]["inspector"].unique())
                with st.popover("🔍 Inspectores", use_container_width=True):
                    insps_sel = st.multiselect("Filtrar:", opc_insps, default=opc_insps)
    
        if df_m.empty or not sups_sel or not insps_sel:
            st.warning("Selecciona un rango y personal válido.")
            st.stop()
    
        df_m = df_m[(df_m["supervisor"].isin(sups_sel)) & (df_m["inspector"].isin(insps_sel))].copy()
        
        # Cálculo de tiempo de recorrido
        def calc_rec(row):
            hi, hr = row.get("hora inicio_parsed"), row.get("hora inicio de recorrido_parsed")
            if not isinstance(hi, datetime.time) or not isinstance(hr, datetime.time): return pd.NaT
            return datetime.datetime.combine(datetime.date.today(), hi) - datetime.datetime.combine(datetime.date.today(), hr)
        df_m["tiempo_recorrido_td"] = df_m.apply(calc_rec, axis=1)
        
        # Marcar efectivas
        valores_efectivos = ["INSPECCIONADA", "INSPECCIONADA CON DEFECTO NO CRITICO", "INSPECCIONADA CON DEFECTO CRITICO", "CERTIFICADA", "CERTIFICADA CON NOVEDAD"]
        df_m["efectiva"] = df_m["cierre"].isin(valores_efectivos)
    
        # --- LÓGICA DE PROMEDIOS DIARIOS ---
        # Convertimos a decimal antes de agrupar para evitar errores de comparación con datetime.time
        df_m["ini_dec_tmp"] = df_m["hora_inicio"].apply(hora_to_decimal)
        df_m["fin_dec_tmp"] = df_m["hora_final"].apply(hora_to_decimal)
    
        # Agrupar por inspector y día para obtener los hitos diarios
        df_daily_hitos = df_m.groupby(["inspector", "fecha"]).agg(
            ini_dec=("ini_dec_tmp", "min"),
            fin_dec=("fin_dec_tmp", "max")
        ).reset_index()
        df_daily_hitos["es_sabado"] = pd.to_datetime(df_daily_hitos["fecha"]).dt.weekday == 5
    
        # KPIs Globales del Rango
        prom_ini_global = df_daily_hitos["ini_dec"].mean()
        # Excluir sábados para el promedio de hora fin
        df_fin_no_sab = df_daily_hitos[~df_daily_hitos["es_sabado"]]
        prom_fin_global = df_fin_no_sab["fin_dec"].mean() if not df_fin_no_sab.empty else None
    
        t_tarea_prom = df_m[df_m["efectiva"] & df_m["tiempo_tarea_td"].notna()]["tiempo_tarea_td"].mean()
        t_rec_prom = df_m["tiempo_recorrido_td"].mean()
        
        total_ord = len(df_m)
        total_eff = df_m["efectiva"].sum()
        perc_eff = round((total_eff / total_ord) * 100, 1) if total_ord > 0 else 0
    
        # Visualización
        col_k, col_v = st.columns([1, 4])
        with col_k:
            st.markdown("### 📈 KPIs Periodo")
            render_kpi("Inicio Prom.", hora_to_string(decimal_to_hora(prom_ini_global)), "⏰")
            render_kpi("Fin Prom. (Lun-Vie)", hora_to_string(decimal_to_hora(prom_fin_global)), "🕒")
            render_kpi("T. Tarea Prom.", td_to_str(t_tarea_prom), "🕓")
            render_kpi("Recorrido Prom.", td_to_str(t_rec_prom), "🚗")
            render_kpi("Total Tareas", total_ord, "📋")
            render_kpi("Efectivas", total_eff, "✅")
            render_kpi("% Efectividad", f"{perc_eff}%", "📈")
    
        with col_v:
            # Resumen por Inspector para la tabla
            resumen_m = df_m.groupby("inspector").apply(lambda x: pd.Series({
                "Dias Trabajados": x["fecha"].nunique(),
                "Total Órdenes": len(x),
                "Efectivas": x["efectiva"].sum(),
                "Efectividad %": round((x["efectiva"].sum() / len(x)) * 100, 1) if len(x) > 0 else 0,
                "Prom. Tarea": td_to_str(x[x["efectiva"]]["tiempo_tarea_td"].mean()),
                "Prom. Recorrido": td_to_str(x["tiempo_recorrido_td"].mean())
            })).reset_index()
    
            # Añadir promedios de horas por inspector
            prom_horas_insp = df_daily_hitos.groupby("inspector").apply(lambda x: pd.Series({
                "Prom. Inicio": hora_to_string(decimal_to_hora(x["ini_dec"].mean())),
                "Prom. Fin (LV)": hora_to_string(decimal_to_hora(x[~x["es_sabado"]]["fin_dec"].mean()))
            })).reset_index()
    
            df_tabla_m = resumen_m.merge(prom_horas_insp, on="inspector")
            
            st.markdown("### 📋 Resumen Consolidado por Inspector")
            st.dataframe(
                df_tabla_m.style.set_properties(**{'text-align': 'center'})
                .set_properties(subset=['inspector'], **{'text-align': 'left'}),
                use_container_width=True, height=500, hide_index=True
            )
    
        # Gráfica de producción mensual
        st.markdown("### 📊 Producción Total en el Periodo")
        df_prod_m = df_m[df_m["efectiva"]].groupby("inspector").size().reset_index(name="Efectivas").sort_values("Efectivas")
        
        if not df_prod_m.empty:
            fig_m = px.bar(df_prod_m, y="inspector", x="Efectivas", orientation="h", text="Efectivas", color="Efectivas", color_continuous_scale="Viridis")
            fig_m.update_layout(height=600, font=dict(size=14))
            st.plotly_chart(fig_m, use_container_width=True)
    
    # =================================================
    # ✅ TAB — SEGUIMIENTO AGENDAS
    # =================================================
    
    @st.fragment
    def render_agendas_alerta_fragment(df_alerta_raw, grupos_validos, columnas_base):
        zonas_sel = []
        with st.expander("Seleccionar Zona"):
            for z in grupos_validos:
                if st.checkbox(z, value=True, key=f"pen_zona_frag_{z}"):
                    zonas_sel.append(z)
        
        df_alerta = df_alerta_raw[df_alerta_raw["grupo"].isin(zonas_sel)] if zonas_sel else df_alerta_raw
    
        if df_alerta.empty:
            st.info("✅ No hay agendas en ALERTA.")
        else:
            st.info("💡 Haz clic en una fila para ver el detalle de la tarea.")
            # Preparamos los datos ordenados para que la selección coincida con el índice
            df_display_alerta = df_alerta[columnas_base].sort_values("fecha de visita").reset_index(drop=True)
            # TABLA RESUMIDA: Inspector, Contrato, Localidad
            cols_tabla = ["inspector", "contrato", "localidad"]
            
            seleccion = st.dataframe(
                df_display_alerta[cols_tabla], 
                use_container_width=True,
                on_select="rerun",
                key="tabla_agendas_alerta_fragment",
                hide_index=True,
                selection_mode="single-row",
                height=600
            )
            st.error(f"🚨 TOTAL ALERTAS: {len(df_alerta)}")
    
            if seleccion.selection.rows:
                idx = seleccion.selection.rows[0]
                
                # LÓGICA ANTI-FANTASMA: Solo mostrar si la selección es nueva
                # Esto evita que el recuadro aparezca solo al filtrar en otras pestañas
                sel_key = f"last_idx_{cols_tabla[0]}_{len(df_display_alerta)}"
                if st.session_state.get(sel_key) != idx:
                    st.session_state[sel_key] = idx
                    fila = df_display_alerta.iloc[idx]
                    
                    fecha_str = fila['fecha de visita'].strftime('%Y-%m-%d') if hasattr(fila['fecha de visita'], 'strftime') else str(fila['fecha de visita'])
                    mostrar_detalle_tarea(
                        fila["contrato"], 
                        fila["detalle de tarea"],
                        direccion=fila["direccion"],
                        fecha=fecha_str,
                        localidad=fila["localidad"],
                        inspector=fila["inspector"]
                    )
    
    with tab_agendas:
        st.markdown("""
            <style>
                .agendas-sidebar {
                    background-color: #2F9331;
                    color: white;
                    padding: 25px 15px 5px 15px;
                    border-radius: 15px 15px 0 0;
                    margin-bottom: 0;
                }
                /* Fusionar el menú con el cuadro azul superior */
                div[data-testid="stRadio"]:has(input[id*="nav_agendas_radio"]) {
                    background-color: #2F9331;
                    padding: 0 15px 25px 15px;
                    border-radius: 0 0 15px 15px;
                    margin-top: -1rem;
                }
                .sidebar-header-age {
                    border-bottom: 1px solid rgba(255,255,255,0.2);
                    margin-bottom: 20px;
                    padding-bottom: 15px;
                }
                /* Estilizar el radio de navegación de agendas */
                div[data-testid="stRadio"] > label {
                    display: none; 
                }
                div[data-testid="stRadio"] div[role="radiogroup"] > label {
                    background-color: rgba(255,255,255,0.05);
                    color: white !important;
                    padding: 10px 15px;
                    border-radius: 8px;
                    margin-bottom: 8px;
                    transition: all 0.3s;
                    border: 1px solid transparent;
                }
                div[data-testid="stRadio"] div[role="radiogroup"] > label:hover {
                    background-color: rgba(255,255,255,0.15);
                }
                div[data-testid="stRadio"] div[role="radiogroup"] > label[data-baseweb="radio"][aria-checked="true"] {
                    background-color: white !important;
                    color: #1e3a8a !important;
                    font-weight: bold;
                }
            </style>
        """, unsafe_allow_html=True)
    
        # ======================================================
        # TÍTULO PRINCIPAL
        st.markdown("## 🗂️ Control agendas")
    
        # ======================================================
        # CARGAR BITÁCORA DESDE GITHUB (FORMA CORRECTA)
        # ======================================================
        # ======================================================
        # CARGAR Y PROCESAR DATOS (CON CACHÉ PARA VELOCIDAD)
        # ======================================================
        token = st.secrets["github"]["token"]
        repo = st.secrets["github"]["repo"]
    
        df = get_processed_agendas_data(repo, token)
        
        if not df.empty:
            ahora_colombia = datetime.datetime.now(ZoneInfo("America/Bogota")).replace(tzinfo=None)
            
            # CÁLCULOS GLOBALES PARA KPIs
            count_final = len(df[df["estado"].str.upper().str.contains("FINALIZAD", na=False)])
            count_prox = len(df[(df["estado"].str.upper().str.contains("ASIGNAD", na=False)) & (df["fecha de ejecucion"].isna()) & (df["fecha de visita"] > ahora_colombia)])
            count_alerta = len(df[(df["estado"].str.upper().str.contains("ASIGNAD", na=False)) & (df["prioridad"].str.upper() == "ALTA") & (df["estado_alerta"] == "ALERTA")])
    
        if not df.empty:
            columnas_base = ["inspector", "contrato", "direccion", "estado", "fecha de visita", "localidad", "detalle de tarea", "estado_alerta"]
            ahora_colombia = datetime.datetime.now(ZoneInfo("America/Bogota")).replace(tzinfo=None)
            grupos_validos = ["INSP-CALDAS", "INSP-RIS"]
    
            # --- LAYOUT CON MENÚ LATERAL ---
            col_nav_age, col_main_age = st.columns([1.2, 4])
    
            with col_nav_age:
                st.markdown(f"""
                    <div class="agendas-sidebar">
                        <div class="sidebar-header-age">
                            <h3 style='color: white; margin:0;'>ESTADO AGENDAS</h3>
                            <p style='color: #cbd5e1; font-size: 0.8rem; margin:0;'>Filtros de seguimiento</p>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
                
                opcion_age = st.radio(
                    "Navegación Agendas",
                    ["✅ Agendas Finalizadas", "⏳ Próximas Agendas", "🚨 Alerta"],
                    key="nav_agendas_radio"
                )
                
                st.markdown("---")
                st.markdown("### 📊 Resumen")
                render_kpi("En Alerta", count_alerta, "🚨")
                render_kpi("Próximas", count_prox, "⏳")
                render_kpi("Finalizadas", count_final, "✅")
    
            with col_main_age:
              if opcion_age == "✅ Agendas Finalizadas":
                st.markdown("### ✅ Agendas finalizadas")
                zonas_sel = []
                with st.expander("Seleccionar Zona"):
                    for z in grupos_validos:
                        if st.checkbox(z, value=True, key=f"fin_zona_{z}"):
                            zonas_sel.append(z)
    
                inicios_sel = []
                with st.expander("Filtrar por inicio de la tarea"):
                    for i in ["INICIO TARDE", "INICIO A TIEMPO"]:
                        if st.checkbox(i, value=True, key=f"fin_inicio_{i}"):
                            inicios_sel.append(i)
    
                df_final = df[df["estado"].str.upper().str.contains("FINALIZAD", na=False)].copy()
                if zonas_sel:
                    df_final = df_final[df_final["grupo"].isin(zonas_sel)]
    
                def evaluar_inicio_tarde(row):
                    if pd.isna(row["fecha de ejecucion"]) or pd.isna(row["fecha de visita"]):
                        return "SIN DATO"
                    limite = row["fecha de visita"] + pd.Timedelta(minutes=20) # Margen de 20 min
                    return "INICIO TARDE" if row["fecha de ejecucion"] > limite else "INICIO A TIEMPO"
    
                df_final["inicio_tarea"] = df_final.apply(evaluar_inicio_tarde, axis=1)
                if inicios_sel:
                    df_final = df_final[df_final["inicio_tarea"].isin(inicios_sel)]
    
                if df_final.empty:
                    st.info("✅ No hay agendas finalizadas con esos filtros.")
                else:
                    st.dataframe(df_final[columnas_base[:-1] + ["inicio_tarea"]].sort_values("fecha de visita"), use_container_width=True, height=600)
    
              elif opcion_age == "⏳ Próximas Agendas":
                st.markdown("### ⏳ Agendas próximas (no iniciadas)")
                df_prox_raw = df[
                    (df["estado"].str.upper().str.contains("ASIGNAD", na=False)) & 
                    (df["fecha de ejecucion"].isna()) & 
                    (df["fecha de visita"] > ahora_colombia)
                ].copy()
                
                zonas_sel = []
                with st.expander("Seleccionar Zona"):
                    for z in grupos_validos:
                        if st.checkbox(z, value=True, key=f"prox_zona_{z}"):
                            zonas_sel.append(z)
                
                df_prox = df_prox_raw[df_prox_raw["grupo"].isin(zonas_sel)] if zonas_sel else df_prox_raw
    
                if df_prox.empty:
                    st.info("✅ No hay agendas próximas.")
                else:
                    st.dataframe(df_prox[columnas_base].sort_values("fecha de visita"), use_container_width=True, height=600)
    
              elif opcion_age == "🚨 Alerta":
                st.markdown("### 🚨 ALERTA")
                df_alerta_raw = df[(df["estado"].str.upper().str.contains("ASIGNAD", na=False)) & (df["prioridad"].str.upper() == "ALTA") & (df["estado_alerta"] == "ALERTA")].copy()
                render_agendas_alerta_fragment(df_alerta_raw, grupos_validos, columnas_base)
        else:
            st.info("No se pudo cargar la bitácora desde GitHub para agendas.")
    # ===================================================
    # ✅ TAB — SEGUIMIENTO ADICIONALES
    # ===================================================
    with tab_adicionales:
        st.subheader("🏭 Seguimiento de Adicionales")
    
        # --- CONFIGURACIÓN DE PERSISTENCIA EN GITHUB ---
        token_ad = st.secrets["github"]["token"]
        repo_ad = st.secrets["github"]["repo"]
        branch_ad = st.secrets["github"].get("branch", "main")
        nombre_archivo_git = "PROGRAMACION.xlsx"
    
        # --- SECCIÓN PARA ACTUALIZAR EL ARCHIVO (Sincronizado con GitHub) ---
        with st.expander("⬆️ Actualizar Base de Datos de Programación"):
            archivo_nuevo = st.file_uploader(
                "Sube el nuevo archivo PROGRAMACION.xlsx para actualizar el dashboard global",
                type=["xlsx", "xls"],
                key="uploader_adicionales_github"
            )
            if st.button("🚀 Guardar y compartir con el equipo", key="btn_subir_adicionales"):
                if archivo_nuevo is not None:
                    contenido_bin = archivo_nuevo.read()
                    contenido_b64_ad = base64.b64encode(contenido_bin).decode("utf-8")
                    
                    url_ad = f"https://api.github.com/repos/{repo_ad}/contents/{nombre_archivo_git}"
                    headers_ad = {"Authorization": f"Bearer {token_ad}", "Accept": "application/vnd.github+json"}
                    
                    # Obtener el SHA actual para permitir el reemplazo (evita conflictos de versión)
                    resp_get = requests.get(url_ad, headers=headers_ad)
                    sha_ad = resp_get.json().get("sha") if resp_get.status_code == 200 else None
                    
                    payload_ad = {
                        "message": "Actualización global de PROGRAMACION.xlsx desde Dashboard",
                        "content": contenido_b64_ad,
                        "branch": branch_ad
                    }
                    if sha_ad: payload_ad["sha"] = sha_ad
                    
                    resp_put = requests.put(url_ad, headers=headers_ad, json=payload_ad)
                    if resp_put.status_code in (200, 201):
                        # --- ACTUALIZAR METADATA DE PROGRAMACIÓN ---
                        info_p = {
                            "ultima_actualizacion": datetime.datetime.now(TZ_UTC).strftime("%Y-%m-%d %H:%M:%S"),
                            "usuario_actualizo": st.session_state.usuario
                        }
                        
                        contenido_info_p_b64 = base64.b64encode(
                            json.dumps(info_p, indent=2).encode("utf-8")
                        ).decode("utf-8")
                        
                        url_info_p = f"https://api.github.com/repos/{repo_ad}/contents/PROGRAMACION_INFO.json"
                        r_info_p = requests.get(url_info_p, headers=headers_ad)
                        sha_info_p = r_info_p.json().get("sha") if r_info_p.status_code == 200 else None
                        
                        payload_info_p = {"message": "Actualización de PROGRAMACION_INFO.json", "content": contenido_info_p_b64, "branch": branch_ad}
                        if sha_info_p: payload_info_p["sha"] = sha_info_p
                        requests.put(url_info_p, headers=headers_ad, json=payload_info_p)
                        
                        st.success("✅ Archivo guardado correctamente en la nube. Ahora todos los usuarios verán esta versión.")
                        fetch_github_excel.clear()
                        process_adicionales_data.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ Error al sincronizar con GitHub: {resp_put.text}")
                else:
                    st.warning("⚠️ Por favor selecciona un archivo antes de intentar guardar.")
    
        # --- CARGA DEL ARCHIVO DESDE GITHUB (Datos compartidos) ---
        df_p, _ = fetch_github_excel(repo_ad, nombre_archivo_git, token_ad, branch_ad)
        
        # Procesamiento cacheado para mayor velocidad
        df_p = process_adicionales_data(df_p)
    
        if not df_p.empty:
            # --- FILTRO DE SEDE (CARGUE) ---
            if "cargue" in df_p.columns:
                sedes_raw = sorted(df_p["cargue"].astype(str).unique().tolist())
                sedes_opciones = ["TODAS"] + sedes_raw
                sedes_sel = st.selectbox("📍 Seleccionar Sede (Cargue):", sedes_opciones, key="filtro_sede_adicionales")
                
                if sedes_sel != "TODAS":
                    df_p = df_p[df_p["cargue"].astype(str) == sedes_sel]
    
            # Selección de columnas y visualización (Fuera del bloque 'if cargue' para mayor robustez)
            cols_req = ["contrato", "nombre_inspector", "direccion barrio", "codigo_tipo_trabajo", "cargue", "dias de asignacion"]
            cols_final = [c for c in cols_req if c in df_p.columns]
            
            def color_semaforo(row):
                # Verificación segura de la existencia de la columna calculada
                if "dias de asignacion" not in row:
                    return [""] * len(row)
                dias = row["dias de asignacion"]
                if dias < 3:
                    return ["background-color: #d4edda; color: #155724"] * len(row)  # Verde
                elif dias == 3:
                    return ["background-color: #fff3cd; color: #856404"] * len(row)  # Amarillo
                else:
                    return ["background-color: #f8d7da; color: #721c24"] * len(row)  # Rojo
    
            st.dataframe(df_p[cols_final].style.apply(color_semaforo, axis=1), use_container_width=True, hide_index=True)
        else:
            st.info("ℹ️ No hay un archivo de programación activo. Utiliza el panel superior para subir 'PROGRAMACION.xlsx'.")
    
    
    # ===================================================
    # ✅ TAB — ÓRDENES ASIGNADAS
    # ===================================================
    with tab_asignadas:
        st.markdown("## 📌 Órdenes ASIGNADAS")
    
        # ===================================================
        # VALIDAR Y CARGAR BITÁCORA LOCAL
        # ===================================================
        archivo_bitacora = "BITACORA.xlsx"
    
        df = df_bitacora_base.copy()
        
        # ===================================================
        # ✅ FILTRAR SOLO GRUPOS PERMITIDOS
        # ===================================================
        if "grupo" in df.columns:
            df["grupo"] = df["grupo"].astype(str).str.upper().str.strip()
    
            # Filtro más flexible para grupos operativos
            df = df[df["grupo"].str.contains("INSP-CALDAS|INSP-RIS", na=False)]
    
            # ===================================================
            # VALIDAR COLUMNAS NECESARIAS
        # ===================================================
        columnas_requeridas = ["inspector", "estado", "prioridad", "grupo"]
        for col in columnas_requeridas:
            if col not in df.columns:
                st.error(f"❌ Falta la columna requerida: {col}")
    
        # ===================================================
        # FILTRAR ÓRDENES EN PROCESO (Asignadas, En Camino, Iniciadas)
        # ===================================================
        estados_carga_regex = "Asignad|En Camino|Iniciada"
        df_asignadas = df[
            df["estado"]
            .astype(str)
            .str.contains(estados_carga_regex, case=False, na=False)
        ].copy()
    
        if df_asignadas.empty:
            st.info("✅ No hay órdenes ASIGNADAS en la bitácora.")
    
        # ===================================================
        # ================= FILTROS =================
        # ===================================================
        st.markdown("### 🔎 Filtros")
    
        # -------- PANEL DE FILTROS TIPO "BOX" - LÓGICA ESTABLE --------
        with st.container(border=True):
            col_f1, col_f2, col_f3, col_f4 = st.columns([0.8, 1.2, 1.2, 1])
    
            # Opciones estables basadas en el conjunto inicial de órdenes
            opc_grupos = sorted(df_asignadas["grupo"].dropna().unique())
            opc_estados = sorted(df_asignadas["estado"].dropna().unique())
            opc_prioridades = sorted(df_asignadas["prioridad"].dropna().unique())
    
            with col_f1:
                grupos_sel = st.pills("📍 Grupo", opc_grupos, selection_mode="multi", default=opc_grupos, key="tab5_grupo_pills")
            with col_f2:
                estados_sel = st.pills("📊 Estado", opc_estados, selection_mode="multi", default=opc_estados, key="tab5_estado_pills")
            with col_f3:
                prioridades_sel = st.pills("⚡ Prioridad", opc_prioridades, selection_mode="multi", default=opc_prioridades, key="tab5_prio_pills")
            with col_f4:
                ver_por = st.segmented_control("📈 Ver por:", ["Prioridad", "Estado"], default="Prioridad", key="tab5_ver_por_seg")
                col_agrupar = ver_por.lower()
    
        # Aplicar todos los filtros al final para evitar reinicios de widgets
        df_finalizados_base = df[df["grupo"].isin(grupos_sel)] if grupos_sel else df
        if grupos_sel: df_asignadas = df_asignadas[df_asignadas["grupo"].isin(grupos_sel)]
        if estados_sel: df_asignadas = df_asignadas[df_asignadas["estado"].isin(estados_sel)]
        if prioridades_sel: df_asignadas = df_asignadas[df_asignadas["prioridad"].isin(prioridades_sel)]
    
        # Identificar inspectores que ya terminaron (Tienen 'Finalizada' y NO tienen carga activa)
        # en los grupos seleccionados para identificar disponibilidad
        insp_con_asig = set(df_finalizados_base[df_finalizados_base["estado"].astype(str).str.contains(estados_carga_regex, case=False, na=False)]["inspector"].unique())
        insp_con_fin = set(df_finalizados_base[df_finalizados_base["estado"].astype(str).str.contains("Finalizad", case=False, na=False)]["inspector"].unique())
        inspectores_finalizados = insp_con_fin - insp_con_asig
    
        if df_asignadas.empty:
            st.warning("⚠️ No hay datos con los filtros seleccionados.")
    
        # ===================================================
        # AGRUPAR POR INSPECTOR Y DIMENSIÓN SELECCIONADA
        # ===================================================
        df_prio = (
            df_asignadas
            .groupby(["inspector", col_agrupar])
            .size()
            .reset_index(name="cantidad")
        )
    
        # Agregar inspectores que ya terminaron su obra (con cantidad 0 para que aparezcan en el eje Y)
        if inspectores_finalizados:
            df_terminados = pd.DataFrame({
                "inspector": list(inspectores_finalizados),
                col_agrupar: "TERMINÓ OBRA",
                "cantidad": 0
            })
            df_prio = pd.concat([df_prio, df_terminados], ignore_index=True)
    
        # Ordenar inspectores por carga total
        orden_inspectores = (
            df_prio.groupby("inspector")["cantidad"].sum()
            .sort_values(ascending=False).index.tolist()
        )
    
        # ===================================================
        # MAPA DE COLORES (Prioridades y Estados)
        # ===================================================
        color_map = {
            # Prioridades
            "Alta": "#dc3545",        # 🔴 rojo
            "Media": "#ffc107",       # 🟡 amarillo
            "Baja": "#7cd992",        # 🟢 verde claro
            "Critica": "#fd7e14",     # 🟠 naranja
            "Prioridad": "#6f4e37",    # 🟤 café
            
            "60 Meses": "#6f42c1",        # 🟣 morado
            "Segunda visita": "#ff8c00",   # 🟠 naranja
    
            # Estados
            "Asignada": "#3498db", "En Camino": "#e67e22", "Iniciada": "#9b59b6",
    
            # Disponibilidad
            "TERMINÓ OBRA": "#28a745"      # 🟢 verde (disponible)
        }
    
        # ===================================================
        # GRÁFICA ACUMULADA
        # ===================================================
        fig = px.bar(
            df_prio,
            y="inspector",
            x="cantidad",
            color=col_agrupar,
            orientation="h",
            category_orders={"inspector": orden_inspectores},
            color_discrete_map=color_map,
            text="cantidad",
            title="Órdenes ASIGNADAS por inspector (según filtros)"
        )
    
        fig.update_traces(
            textposition="inside",
            textfont_size=16
        )
    
        fig.update_layout(
            barmode="stack",
            xaxis_title="Cantidad de órdenes ASIGNADAS",
            yaxis_title="Inspector",
            legend_title=ver_por,
            height=700
        )
    
        st.plotly_chart(fig, use_container_width=True)
    
    
    # ===================================================
    # ✅ TAB — INVENTARIO V2.
    # ===================================================
with tab_inv_v2:
    st.markdown("""
        <style>
            #inv-v2-container {
                background-color: transparent;
                font-family: 'Inter', sans-serif;
            }
            #inv-v2-container .filter-card {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                padding: 24px;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05), 0 2px 4px -1px rgba(0, 0, 0, 0.06);
                margin-bottom: 24px;
            }
            #inv-v2-container .filter-title {
                color: #0f172a;
                text-align: center;
                font-size: 1.25rem;
                font-weight: 700;
                margin-bottom: 20px;
                text-transform: uppercase;
                letter-spacing: 0.05em;
            }
            #inv-v2-container .results-card {
                background-color: #ffffff;
                border: 1px solid #e2e8f0;
                border-radius: 12px;
                padding: 20px;
                box-shadow: 0 4px 6px -1px rgba(0, 0, 0, 0.05);
            }
            #inv-v2-container .section-label {
                font-weight: 700;
                color: #1e293b;
                margin-bottom: 16px;
                font-size: 1.1rem;
                border-left: 4px solid #3b82f6;
                padding-left: 12px;
            }
        </style>
        <div id="inv-v2-container">
    """, unsafe_allow_html=True)
    
    CATALOGO_DEFAULT = {
        "EPPs": {
            "Monogafas":  {"tallas": False},
            "Guantes":    {"tallas": False},
            "Piernera":   {"tallas": False},
        },
        "Dotación": {
            "Botas":    {"tallas": True, "opciones_talla": ["36","37","38","39","40","41","42","43","44","45","46"]},
            "Camisa":   {"tallas": True, "opciones_talla": ["XS","S","M","L","XL","XXL"]},
            "Pantalón": {"tallas": True, "opciones_talla": ["28","30","32","34","36","38","40"]},
            "Chaleco":  {"tallas": True, "opciones_talla": ["XS","S","M","L","XL","XXL"]},
        },
        "Papelería": {
            "Isométricos (paquete x200)": {"tallas": False},
            "Stickers":                   {"tallas": False},
            "Papelería general":          {"tallas": False},
        },
        "Herramientas": {
            "Cepo":           {"tallas": False},
            "Llaves de cepo": {"tallas": False},
        },
    }

    SEDES_INV = ["CALDAS", "RISARALDA"]
    RESPONSABLES_INV = ["CRISTIAN CHICA", "JANIER", "JENNY", "CAMILA", "ANDRES", "DANNY"]
    inv_token = st.secrets["github"]["token"]
    inv_repo  = st.secrets["github"]["repo"]
    inv_branch = st.secrets["github"].get("branch", "main")

    movimientos, _ = fetch_github_json(inv_repo, "INVENTARIO_V2.json", inv_token)
    catalogo, _    = fetch_github_json(inv_repo, "CATALOGO_V2.json", inv_token)

    if not isinstance(movimientos, list): movimientos = []
    if not isinstance(catalogo, dict) or not catalogo: catalogo = CATALOGO_DEFAULT.copy()

    def calcular_stock(movs, sede):
        res = {}
        for m in [x for x in movs if x["sede"] == sede]:
            k = f"{m['categoria']}|{m['item']}|{m.get('talla') or 'N/A'}"
            res.setdefault(k, {"categoria": m["categoria"], "item": m["item"], "talla": m.get("talla") or "N/A", "ent": 0, "sal": 0})
            if m["tipo"] == "ENTRADA": res[k]["ent"] += m["cantidad"]
            else: res[k]["sal"] += m["cantidad"]
        
        return pd.DataFrame([
            {"Categoría": v["categoria"], "Ítem": v["item"], "Talla": v["talla"], 
             "Entradas": v["ent"], "Salidas": v["sal"], "Stock": v["ent"]-v["sal"]} 
            for v in res.values()
        ])

    st.markdown('<div class="filter-card"><div class="filter-title">Filtros de Inventario</div>', unsafe_allow_html=True)
    col_f1, _ = st.columns([1, 2])
    with col_f1:
        sede_global = st.selectbox("📍 Seleccionar Sede Global:", SEDES_INV, key="inv_sede_global_main")
    st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="results-card">', unsafe_allow_html=True)
    t_stock, t_ent, t_sal, t_hist, t_conf = st.tabs(["📊 Stock Actual", "➕ Registrar Entrada", "➖ Registrar Salida", "📜 Historial", "⚙️ Configuración"])
    
    with t_stock:
        df_stock = calcular_stock(movimientos, sede_global)
        if not df_stock.empty:
            st.dataframe(df_stock.sort_values(["Categoría", "Stock"]), use_container_width=True, hide_index=True)
        else:
            st.info("No hay inventario registrado en esta sede.")
            
    with t_ent:
        with st.container():
            st.markdown('<p class="section-label">📥 Registro de Entrada</p>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            f_sede = c1.selectbox("Sede Destino", SEDES_INV, key="ent_sede")
            f_resp = c2.selectbox("Responsable Recibo", RESPONSABLES_INV, key="ent_resp")
            f_fecha = c3.date_input("Fecha Recibo", key="ent_fecha")

            st.markdown("---")
            c4, c5, c6, c7 = st.columns([1.5, 1.5, 1, 1])
            f_cat = c4.selectbox("Categoría", list(catalogo.keys()), key="ent_cat")
            f_item = c5.selectbox("Producto", list(catalogo[f_cat].keys()), key="ent_item")
            
            opciones_talla = catalogo[f_cat][f_item].get("opciones_talla", [])
            f_talla = c6.selectbox("Talla", opciones_talla if opciones_talla else ["N/A"], key="ent_talla")
            f_cant = c7.number_input("Cantidad", min_value=1, step=1, key="ent_cant")
            f_obs = st.text_area("Observaciones / Remisión", key="ent_obs")

            if st.button("💾 Guardar Entrada", type="primary", use_container_width=True, key="btn_save_ent"):
                nuevo = {
                    "tipo": "ENTRADA", "fecha": str(f_fecha), "sede": f_sede,
                    "responsable": f_resp, "categoria": f_cat, "item": f_item,
                    "talla": f_talla if f_talla != "N/A" else None, "cantidad": f_cant,
                    "observacion": f_obs, "timestamp": datetime.datetime.now(TZ_CO).strftime("%Y-%m-%d %H:%M:%S")
                }
                movimientos.append(nuevo)
                save_github_json(inv_repo, "INVENTARIO_V2.json", inv_token, movimientos, f"Entrada {f_item}")
                st.success("Ingreso registrado")
                st.rerun()

    with t_sal:
        with st.container():
            st.markdown('<p class="section-label">📤 Registro de Salida</p>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            f_sede = c1.selectbox("Sede Origen", SEDES_INV, key="sal_sede")
            f_resp = c2.selectbox("Entrega", RESPONSABLES_INV, key="sal_resp")
            f_insp = c3.selectbox("Recibe (Inspector)", inspectores_lista, key="sal_insp")

            st.markdown("---")
            c4, c5, c6, c7 = st.columns([1.5, 1.5, 1, 1])
            f_cat = c4.selectbox("Categoría", list(catalogo.keys()), key="sal_cat")
            f_item = c5.selectbox("Producto", list(catalogo[f_cat].keys()), key="sal_item")
            
            opciones_talla = catalogo[f_cat][f_item].get("opciones_talla", [])
            f_talla = c6.selectbox("Talla", opciones_talla if opciones_talla else ["N/A"], key="sal_talla")
            f_cant = c7.number_input("Cantidad a entregar", min_value=1, step=1, key="sal_cant")
            
            if st.button("✅ Procesar Salida", type="primary", use_container_width=True, key="btn_save_sal"):
                df_stock = calcular_stock(movimientos, f_sede)
                talla_val = f_talla if f_talla != "N/A" else "N/A"
                match = df_stock[(df_stock["Categoría"] == f_cat) & (df_stock["Ítem"] == f_item) & (df_stock["Talla"] == talla_val)]
                disponible = match["Stock"].iloc[0] if not match.empty else 0
                
                if f_cant > disponible:
                    st.error(f"❌ No hay suficiente stock. Disponible: {disponible}")
                else:
                    nuevo = {
                        "tipo": "SALIDA", "fecha": str(datetime.date.today()), "sede": f_sede,
                        "responsable": f_resp, "inspector": f_insp, "categoria": f_cat,
                        "item": f_item, "talla": f_talla if f_talla != "N/A" else None,
                        "cantidad": f_cant, "timestamp": datetime.datetime.now(TZ_CO).strftime("%Y-%m-%d %H:%M:%S")
                    }
                    movimientos.append(nuevo)
                    save_github_json(inv_repo, "INVENTARIO_V2.json", inv_token, movimientos, f"Salida {f_item}")
                    st.success("Entrega registrada")
                    st.rerun()

    with t_hist:
        if movimientos:
            df_h = pd.DataFrame(movimientos)
            col_h1, col_h2, col_h3 = st.columns(3)
            filter_sede = col_h1.selectbox("📍 Sede", ["TODAS"] + SEDES_INV, key="hist_f_sede")
            filter_tipo = col_h2.selectbox("Movimiento", ["TODOS", "ENTRADA", "SALIDA"], key="hist_f_tipo")
            filter_cat  = col_h3.selectbox("Categoría", ["TODAS"] + list(catalogo.keys()), key="hist_f_cat")

            filtered_df_h = df_h.copy()
            if filter_sede != "TODAS": filtered_df_h = filtered_df_h[filtered_df_h["sede"] == filter_sede]
            if filter_tipo != "TODOS": filtered_df_h = filtered_df_h[filtered_df_h["tipo"] == filter_tipo]
            if filter_cat != "TODAS":  filtered_df_h = filtered_df_h[filtered_df_h["categoria"] == filter_cat]

            st.dataframe(filtered_df_h.sort_values("timestamp", ascending=False), use_container_width=True, hide_index=True)
        else:
            st.info("No hay historial.")

    with t_conf:
        st.markdown('<p class="section-label">⚙️ Configuración del Maestro</p>', unsafe_allow_html=True)
        with st.expander("Ver Catálogo Actual"):
            st.json(catalogo)
            
    st.markdown('</div> <!-- Cierra results-card -->', unsafe_allow_html=True)
    st.markdown('</div> <!-- Cierra container principal -->', unsafe_allow_html=True)

# ✅ TAB — SST
# ===================================================
with tab_sst:
    st.subheader("🦺 SST - Seguimiento Preoperacionales, Operacionales y Ausentismo")

    # Asegurarse de que df_bitacora_base esté disponible y tenga las columnas necesarias
    required_cols = ["contrato", "tipo de trabajo", "inspector", "hora final_parsed", "tiempo de tarea"]
    if df_bitacora_base is None or df_bitacora_base.empty:
        st.info("ℹ️ No hay datos en la bitácora para mostrar en SST.")
    elif not all(col in df_bitacora_base.columns for col in required_cols):
        missing_cols = [col for col in required_cols if col not in df_bitacora_base.columns]
        st.error(f"❌ Faltan columnas requeridas ({', '.join(missing_cols)}) en la bitácora para la sección SST.")
    else:
        # --- PANELsstv3 DE FILTROS PARA SST (Sincronizado con Seguimiento Diario y Asignadas) ---
        with st.container(border=True):
            col_f1, col_f2 = st.columns([1, 2.4])
            
            with col_f1:
                fechas_validas_sst = sorted(df_bitacora_base["fecha"].dropna().unique())
                fecha_sel_sst = st.selectbox("📅 Fecha Consulta SST:", fechas_validas_sst, key="sst_fecha_sel")
                df_base_sst = df_bitacora_base[df_bitacora_base["fecha"] == fecha_sel_sst].copy()

            with col_f2:
                opc_sups_sst = sorted(df_base_sst["supervisor"].unique())
                supervisores_sel_sst = st.pills("👥 Supervisores:", opc_sups_sst, selection_mode="multi", default=opc_sups_sst, key="sst_sup_pills")

        if not supervisores_sel_sst:
            st.warning("⚠️ Selecciona al menos un supervisor para ver los datos de SST.")
            st.stop()
            
        # Aplicar filtros de Supervisor sobre la fecha seleccionada
        df_sst_filtered = df_base_sst[df_base_sst["supervisor"].isin(supervisores_sel_sst)].copy()

        # Filtrar por el contrato específico "OFM-2025-014, EJE" sobre los datos ya filtrados
        df_eje_contract = df_sst_filtered[
            df_sst_filtered["contrato"].astype(str).str.upper().str.strip() == "OFM-2025-014, EJE"
        ].copy()

        if df_eje_contract.empty:
            st.info("ℹ️ No se encontraron registros para el contrato 'OFM-2025-014, EJE'.")
        else:
            # --- Extraer datos de Preoperacional ---
            df_preoperacional = df_eje_contract[
                df_eje_contract["tipo de trabajo"].astype(str).str.upper().str.strip() == "PREOPERACIONAL - 2025 - EJE"
            ].copy()
            
            df_preoperacional_agg = df_preoperacional.groupby("inspector").agg(
                HORA_PREOPERACIONAL=("hora final_parsed", lambda x: x.iloc[0] if not x.empty else pd.NaT)
            ).reset_index()

            # --- Extraer datos de Operacional Final ---
            df_operacional_final = df_eje_contract[
                df_eje_contract["tipo de trabajo"].astype(str).str.upper().str.strip() == "OPERACIONAL FINAL - EJE"
            ].copy()
            
            df_operacional_final_agg = df_operacional_final.groupby("inspector").agg(
                HORA_OPERACIONAL_FINAL=("hora final_parsed", lambda x: x.iloc[0] if not x.empty else pd.NaT)
            ).reset_index()

            # --- Extraer datos de Ausentismo (Almuerzo) ---
            df_ausentismo = df_eje_contract[
                df_eje_contract["tipo de trabajo"].astype(str).str.upper().str.strip() == "AUSENTISMO"
            ].copy()
            
            df_ausentismo_agg = df_ausentismo.groupby("inspector").agg(
                TIEMPO_AUSENTISMO=("tiempo de tarea", lambda x: str(x.iloc[0]) if not x.empty else "00:00:00")
            ).reset_index()

            # Obtener todos los inspectores únicos de ambos conjuntos de datos
            all_inspectors_in_contract = pd.concat([
                df_preoperacional_agg["inspector"],
                df_operacional_final_agg["inspector"],
                df_ausentismo_agg["inspector"]
            ]).unique()

            df_result = pd.DataFrame({"inspector": all_inspectors_in_contract})

            # Unir los datos
            df_result = df_result.merge(df_preoperacional_agg, on="inspector", how="left")
            df_result = df_result.merge(df_operacional_final_agg, on="inspector", how="left")
            df_result = df_result.merge(df_ausentismo_agg, on="inspector", how="left")

            # Formatear las horas y manejar valores faltantes
            df_result["HORA PREOPERACIONAL"] = df_result["HORA_PREOPERACIONAL"].apply(
                lambda x: x.strftime("%H:%M") if pd.notna(x) else "SIN PREOPERACIONAL"
            )
            df_result["HORA OPERACIONAL FINAL"] = df_result["HORA_OPERACIONAL_FINAL"].apply(
                lambda x: x.strftime("%H:%M") if pd.notna(x) else "SIN OPERACIONAL FINAL"
            )
            
            df_result["AUSENTISMO"] = df_result["TIEMPO_AUSENTISMO"].fillna("00:00:00").apply(
                lambda x: "SIN AUSENTISMO" if str(x) == "00:00:00" else str(x)
            )

            # Renombrar columnas para la visualización final
            df_result = df_result.rename(columns={"inspector": "INSPECTOR"})

            # --- Estilos Condicionales ---
            def style_sst(row):
                styles = [''] * len(row)
                
                # Rojo para Preoperacional y Operacional Final faltantes
                if row["HORA PREOPERACIONAL"] == "SIN PREOPERACIONAL":
                    styles[row.index.get_loc("HORA PREOPERACIONAL")] = 'background-color: #f8d7da; color: #721c24'
                if row["HORA OPERACIONAL FINAL"] == "SIN OPERACIONAL FINAL":
                    styles[row.index.get_loc("HORA OPERACIONAL FINAL")] = 'background-color: #f8d7da; color: #721c24'
                
                # Rojo para Ausentismo (faltante, <30m o >1h5m)
                aus = row["AUSENTISMO"]
                idx_aus = row.index.get_loc("AUSENTISMO")
                if aus == "SIN AUSENTISMO":
                    styles[idx_aus] = 'background-color: #f8d7da; color: #721c24'
                else:
                    try:
                        td = pd.to_timedelta(aus)
                        if td < pd.Timedelta(minutes=30) or td > pd.Timedelta(hours=1, minutes=5):
                            styles[idx_aus] = 'background-color: #f8d7da; color: #721c24'
                    except:
                        pass
                return styles

            # Mostrar la tabla final
            if not df_result.empty:
                st.markdown(f"### 📋 Control de Horarios y Ausentismo")
                st.dataframe(
                    df_result[["INSPECTOR", "HORA PREOPERACIONAL", "HORA OPERACIONAL FINAL", "AUSENTISMO"]].style.apply(style_sst, axis=1),
                    use_container_width=True,
                    hide_index=True
                )
            else:
                st.info("ℹ️ No se encontraron registros para el contrato 'OFM-2025-014, EJE' en esta fecha.")

# ===================================================
# ✅ TAB — SUBIR ARCHIVOS
# ===================================================
with tab_subir:
    st.subheader("📈 Administración de Archivos")

    with st.expander("📂 Cargar Bitácora Operativa (BITACORA.xlsx)", expanded=True):
        st.info("Este proceso reemplaza la base de datos principal utilizada en el Seguimiento Diario y Agendas.")
        
        archivo_bit = st.file_uploader("Selecciona BITACORA.xlsx", type=["xlsx", "xls"], key="uploader_bit_global")
        
        if st.button("🚀 Actualizar Bitácora", use_container_width=True, key="btn_bit_global"):
            if archivo_bit:
                with st.spinner("Sincronizando con GitHub..."):
                    token = st.secrets["github"]["token"]
                    repo = st.secrets["github"]["repo"]
                    branch = st.secrets["github"].get("branch", "main")
                    
                    # 1. Subir el archivo Excel a GitHub
                    content_bin = archivo_bit.read()
                    content_b64 = base64.b64encode(content_bin).decode("utf-8")
                    
                    url_git = f"https://api.github.com/repos/{repo}/contents/BITACORA.xlsx"
                    headers = {"Authorization": f"Bearer {token}", "Accept": "application/vnd.github+json"}
                    
                    r_get = requests.get(url_git, headers=headers)
                    sha = r_get.json().get("sha") if r_get.status_code == 200 else None
                    
                    payload = {
                        "message": f"Actualización de BITACORA.xlsx por {st.session_state.usuario}",
                        "content": content_b64,
                        "branch": branch
                    }
                    if sha: payload["sha"] = sha
                    
                    r_put = requests.put(url_git, headers=headers, json=payload)
                    
                    if r_put.status_code in (200, 201):
                        # 2. Actualizar metadata de la bitácora (BITACORA_INFO.json)
                        info_data = {
                            "ultima_actualizacion": datetime.datetime.now(TZ_UTC).strftime("%Y-%m-%d %H:%M:%S"),
                            "usuario_actualizo": st.session_state.usuario
                        }
                        save_github_json(repo, "BITACORA_INFO.json", token, info_data, "Update BITACORA_INFO.json", branch)
                        
                        st.success("✅ Bitácora actualizada y sincronizada para todos los usuarios.")
                        fetch_github_excel.clear()
                        load_local_bitacora.clear()
                        extract_excel_links.clear()
                        st.rerun()
                    else:
                        st.error(f"❌ Error al subir: {r_put.text}")
            else:
                st.warning("⚠️ Selecciona un archivo Excel antes de subir.")
